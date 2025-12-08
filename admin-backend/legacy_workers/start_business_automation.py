#!/usr/bin/env python3
"""
完整業務自動化啟動腳本
整合功能：
- 自動建群（業務驅動）
- 聊天進度追蹤
- 根據進度自動邀請真實用戶
- 劇本系統驅動的智能對話
- 紅包自動化
"""

import os
import sys
import asyncio
import logging
import yaml
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

# 設置日誌
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s"
)
logger = logging.getLogger("BusinessAutomation")

# 導入依賴
try:
    from telethon import TelegramClient, events
    from telethon.tl.types import Message, User, Chat, Channel
except ImportError:
    print("錯誤：請安裝 telethon: pip install telethon")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("錯誤：請安裝 openpyxl: pip install openpyxl")
    sys.exit(1)

try:
    import httpx
except ImportError:
    print("錯誤：請安裝 httpx: pip install httpx")
    sys.exit(1)

# 導入業務自動化模組
from worker_business_automation import (
    FullBusinessAutomation,
    ChatStage,
    UserEngagementLevel
)
from worker_auto_redpacket import RedPacketConfig, GroupInteractionManager
from worker_group_manager import load_accounts_with_validation


# ==================== 配置 ====================

# 基本配置
SESSIONS_DIR = os.getenv("SESSIONS_DIR", "./sessions")
SCRIPTS_DIR = os.getenv("SCRIPTS_DIR", "./scripts")

# 紅包 API 配置
REDPACKET_API_URL = os.getenv("REDPACKET_API_URL", "https://api.usdt2026.cc")
REDPACKET_API_KEY = os.getenv("REDPACKET_API_KEY", "test-key-2024")

# 業務配置
AUTO_CREATE_GROUP = os.getenv("AUTO_CREATE_GROUP", "true").lower() == "true"
GROUP_NAME = os.getenv("GROUP_NAME", "")
SCRIPT_ID = os.getenv("SCRIPT_ID", "红包游戏陪玩剧本")

# 用戶邀請配置
INVITE_USERS_FILE = os.getenv("INVITE_USERS_FILE", "")  # 待邀請用戶列表文件
AUTO_INVITE = os.getenv("AUTO_INVITE", "false").lower() == "true"

# 自動化設置
AUTO_GRAB = os.getenv("AUTO_GRAB", "true").lower() == "true"
AUTO_SEND = os.getenv("AUTO_SEND", "false").lower() == "true"
AUTO_CHAT = os.getenv("AUTO_CHAT", "true").lower() == "true"


# ==================== 劇本載入 ====================

def load_script(script_path: str) -> Optional[dict]:
    """載入劇本文件"""
    try:
        with open(script_path, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f)
    except Exception as e:
        logger.error(f"載入劇本失敗: {e}")
        return None


def find_script_file(script_id: str) -> Optional[str]:
    """查找劇本文件"""
    # 搜索路徑
    search_paths = [
        Path(SCRIPTS_DIR),
        Path("ai_models/group_scripts"),
        Path("../ai_models/group_scripts"),
    ]
    
    for base_path in search_paths:
        if not base_path.exists():
            continue
        
        # 嘗試不同的文件名格式
        patterns = [
            f"{script_id}.yaml",
            f"{script_id}.yml",
        ]
        
        for pattern in patterns:
            script_path = base_path / pattern
            if script_path.exists():
                return str(script_path)
        
        # 搜索包含 script_id 的文件
        for yaml_file in base_path.glob("*.yaml"):
            if script_id in yaml_file.stem:
                return str(yaml_file)
    
    return None


# ==================== 用戶池載入 ====================

def load_invite_users(file_path: str) -> List[dict]:
    """載入待邀請用戶列表"""
    users = []
    
    if not file_path or not Path(file_path).exists():
        return users
    
    try:
        if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
            # Excel 格式
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            headers = [str(cell.value).lower().strip() if cell.value else "" for cell in ws[1]]
            col_map = {}
            
            for idx, header in enumerate(headers):
                if header in ["user_id", "userid", "tg_id", "id"]:
                    col_map["user_id"] = idx
                elif header in ["username", "用户名"]:
                    col_map["username"] = idx
                elif header in ["phone", "手机", "電話"]:
                    col_map["phone"] = idx
                elif header in ["name", "名字"]:
                    col_map["name"] = idx
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                
                user = {}
                if "user_id" in col_map and row[col_map["user_id"]]:
                    user["user_id"] = int(row[col_map["user_id"]])
                if "username" in col_map and row[col_map["username"]]:
                    user["username"] = str(row[col_map["username"]])
                if "phone" in col_map and row[col_map["phone"]]:
                    user["phone"] = str(row[col_map["phone"]])
                if "name" in col_map and row[col_map["name"]]:
                    user["first_name"] = str(row[col_map["name"]])
                
                if user.get("user_id") or user.get("username") or user.get("phone"):
                    users.append(user)
        
        elif file_path.endswith('.txt'):
            # 文本格式（每行一個 user_id 或 username）
            with open(file_path, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue
                    
                    if line.isdigit():
                        users.append({"user_id": int(line)})
                    elif line.startswith('@'):
                        users.append({"username": line[1:]})
                    else:
                        users.append({"username": line})
        
        logger.info(f"載入了 {len(users)} 個待邀請用戶")
        
    except Exception as e:
        logger.error(f"載入用戶列表失敗: {e}")
    
    return users


# ==================== Excel 文件查找 ====================

def find_excel_file(sessions_dir: str) -> Optional[str]:
    """查找 Excel 配置文件"""
    sessions_path = Path(sessions_dir)
    
    priority_names = ["accounts.xlsx", "config.xlsx", "帳號.xlsx", "账号.xlsx"]
    for name in priority_names:
        excel_path = sessions_path / name
        if excel_path.exists():
            return str(excel_path)
    
    xlsx_files = list(sessions_path.glob("*.xlsx"))
    if xlsx_files:
        return str(xlsx_files[0])
    
    return None


# ==================== 帳號連接 ====================

async def connect_account(
    session_file: str,
    account_config: Dict[str, Any]
) -> Optional[tuple]:
    """連接單個帳號"""
    api_id = account_config.get("api_id")
    api_hash = account_config.get("api_hash")
    phone = account_config.get("phone", "未知")
    
    if not api_id or not api_hash:
        logger.error(f"帳號 {phone} 缺少 API 憑證")
        return None
    
    try:
        client = TelegramClient(session_file, int(api_id), api_hash)
        await client.connect()
        
        if not await client.is_user_authorized():
            logger.error(f"帳號 {phone} 未授權")
            await client.disconnect()
            return None
        
        me = await client.get_me()
        user_id = me.id
        
        account_config["user_id"] = user_id
        account_config["username"] = me.username
        account_config["name"] = f"{me.first_name or ''} {me.last_name or ''}".strip()
        
        logger.info(f"✅ 帳號已連接: {me.username or phone} (ID: {user_id})")
        
        return (client, user_id, account_config)
        
    except Exception as e:
        logger.error(f"連接帳號 {phone} 失敗: {e}")
        return None


def scan_session_files(sessions_dir: str) -> List[str]:
    """掃描 session 文件"""
    session_files = []
    sessions_path = Path(sessions_dir)
    
    if not sessions_path.exists():
        return session_files
    
    for f in sessions_path.glob("*.session"):
        session_files.append(str(f))
    
    return session_files


def match_session_with_config(
    session_file: str,
    accounts: List[Dict[str, Any]]
) -> Optional[Dict[str, Any]]:
    """匹配 session 文件與帳號配置"""
    session_name = Path(session_file).stem
    
    for account in accounts:
        phone = account.get("phone", "")
        if phone and (phone in session_name or session_name in phone):
            if not account.get("api_id") or not account.get("api_hash"):
                logger.error(f"帳號 {phone} 缺少獨立的 API 憑證")
                return None
            return account
    
    return None


# ==================== 劇本驅動的消息處理 ====================

class ScriptDrivenMessageHandler:
    """劇本驅動的消息處理器"""
    
    def __init__(
        self,
        automation: FullBusinessAutomation,
        script_data: dict,
        redpacket_config: RedPacketConfig
    ):
        self.automation = automation
        self.script_data = script_data
        self.redpacket_config = redpacket_config
        self.redpacket_managers: Dict[int, GroupInteractionManager] = {}
        
        # 解析劇本中的角色
        self.roles = {}
        if script_data and "metadata" in script_data:
            for role in script_data["metadata"].get("roles", []):
                self.roles[role["id"]] = role
        
        # 解析場景
        self.scenes = script_data.get("scenes", {}) if script_data else {}
    
    def get_response_for_trigger(
        self,
        group_id: int,
        trigger_type: str,
        message_text: str = ""
    ) -> Optional[dict]:
        """根據觸發條件獲取回復"""
        session = self.automation.progress_tracker.get_session(group_id)
        if not session:
            return None
        
        # 根據階段選擇場景
        stage = session.current_stage
        
        scene_mapping = {
            ChatStage.INITIAL: ["scene1_welcome", "scene2_casual_chat"],
            ChatStage.WARMING_UP: ["scene2_casual_chat", "scene8_random_topics"],
            ChatStage.READY_FOR_USERS: ["scene1_welcome", "scene2_casual_chat"],
            ChatStage.USERS_JOINED: ["scene1_welcome", "scene2_casual_chat"],
            ChatStage.ENGAGING: ["scene2_casual_chat", "scene3_introduce_game"],
            ChatStage.GAME_INTRODUCED: ["scene3_introduce_game", "scene4_game_playing"],
            ChatStage.GAME_PLAYING: ["scene4_game_playing", "scene5_game_result"],
            ChatStage.CONVERSION: ["scene5_game_result", "scene6_continue_chat"],
        }
        
        available_scenes = scene_mapping.get(stage, ["scene2_casual_chat"])
        
        # 查找匹配的回復
        for scene_id in available_scenes:
            scene = self.scenes.get(scene_id)
            if not scene:
                continue
            
            responses = scene.get("responses", [])
            if responses:
                # 根據概率選擇回復
                import random
                for resp in responses:
                    if random.random() < resp.get("probability", 0.5):
                        return {
                            "scene_id": scene_id,
                            "response": resp,
                            "speaker": resp.get("speaker", ""),
                            "text": resp.get("template", ""),
                            "delay_min": resp.get("delay_min", 5),
                            "delay_max": resp.get("delay_max", 30)
                        }
        
        return None
    
    async def handle_new_message(
        self,
        client: TelegramClient,
        user_id: int,
        event: events.NewMessage.Event
    ):
        """處理新消息"""
        message = event.message
        chat = await event.get_chat()
        
        # 只處理群消息
        if not isinstance(chat, (Chat, Channel)):
            return
        
        group_id = chat.id
        text = message.text or ""
        sender_id = message.sender_id
        
        # 忽略自己的消息
        if sender_id == user_id:
            return
        
        # 記錄消息到進度追蹤器
        session = self.automation.progress_tracker.get_session(group_id)
        if session:
            is_ai = sender_id in session.ai_members
            self.automation.progress_tracker.record_message(
                group_id, sender_id, text, is_ai=is_ai
            )
        
        # 處理紅包
        if self._is_redpacket_message(text):
            await self._handle_redpacket(client, user_id, group_id, text)
            return
        
        # 獲取劇本回復
        response_data = self.get_response_for_trigger(
            group_id,
            "message",
            text
        )
        
        if response_data:
            # 添加延遲
            import random
            delay = random.uniform(
                response_data["delay_min"],
                response_data["delay_max"]
            )
            await asyncio.sleep(delay)
            
            # 發送回復
            try:
                await client.send_message(group_id, response_data["text"])
                logger.info(f"[{user_id}] 發送回復: {response_data['text'][:30]}...")
            except Exception as e:
                logger.error(f"發送回復失敗: {e}")
    
    def _is_redpacket_message(self, text: str) -> bool:
        """判斷是否是紅包消息"""
        keywords = ["紅包", "红包", "🧧", "💰", "packet", "hongbao"]
        return any(kw in text.lower() for kw in keywords)
    
    async def _handle_redpacket(
        self,
        client: TelegramClient,
        user_id: int,
        group_id: int,
        text: str
    ):
        """處理紅包消息"""
        # 獲取或創建紅包管理器
        if user_id not in self.redpacket_managers:
            self.redpacket_managers[user_id] = GroupInteractionManager(
                client, user_id, self.redpacket_config
            )
        
        manager = self.redpacket_managers[user_id]
        
        # 提取紅包 UUID
        packet_uuid = manager.extract_packet_uuid(text)
        if packet_uuid and self.redpacket_config.auto_grab:
            claimed = await manager.claim_redpacket(packet_uuid)
            
            if claimed:
                # 記錄紅包活動
                self.automation.progress_tracker.record_redpacket_activity(
                    group_id, user_id, "claim", claimed
                )


# ==================== 主程序 ====================

async def main():
    """主函數"""
    
    print("=" * 70)
    print("  🚀 完整業務自動化系統")
    print("  功能：自動建群 | 聊天進度追蹤 | 智能邀請用戶 | 劇本驅動對話")
    print("=" * 70)
    print()
    
    # 查找 Excel 文件
    excel_file = find_excel_file(SESSIONS_DIR)
    if not excel_file:
        logger.error(f"在 {SESSIONS_DIR} 目錄下找不到 Excel 配置文件！")
        return
    
    logger.info(f"使用 Excel 配置: {excel_file}")
    
    # 載入帳號配置
    accounts = load_accounts_with_validation(excel_file)
    if not accounts:
        logger.error("沒有有效的帳號配置！")
        return
    
    print(f"📋 載入了 {len(accounts)} 個帳號配置")
    
    # 掃描 session 文件
    session_files = scan_session_files(SESSIONS_DIR)
    if not session_files:
        logger.error("沒有找到任何 session 文件！")
        return
    
    # 連接所有帳號
    connected = []
    clients_dict = {}
    
    for session_file in session_files:
        account_config = match_session_with_config(session_file, accounts)
        if account_config:
            result = await connect_account(session_file, account_config)
            if result:
                client, user_id, config = result
                connected.append((client, user_id, config))
                clients_dict[user_id] = client
    
    if not connected:
        logger.error("沒有成功連接任何帳號！")
        return
    
    print(f"\n✅ 成功連接 {len(connected)} 個帳號")
    
    # 創建業務自動化系統
    automation = FullBusinessAutomation()
    
    for client, user_id, config in connected:
        automation.add_client(user_id, client)
    
    # 載入劇本
    script_data = None
    if SCRIPT_ID:
        script_path = find_script_file(SCRIPT_ID)
        if script_path:
            script_data = load_script(script_path)
            if script_data:
                logger.info(f"✅ 載入劇本: {script_data.get('script_id', SCRIPT_ID)}")
    
    # 載入待邀請用戶
    if INVITE_USERS_FILE and AUTO_INVITE:
        invite_users = load_invite_users(INVITE_USERS_FILE)
        if invite_users:
            automation.add_users_to_invite_pool(invite_users)
            logger.info(f"✅ 載入 {len(invite_users)} 個待邀請用戶")
    
    # 創建紅包配置
    redpacket_config = RedPacketConfig()
    redpacket_config.api_url = REDPACKET_API_URL
    redpacket_config.api_key = REDPACKET_API_KEY
    redpacket_config.auto_grab = AUTO_GRAB
    redpacket_config.auto_send = AUTO_SEND
    redpacket_config.auto_chat = AUTO_CHAT
    
    # 創建消息處理器
    message_handler = ScriptDrivenMessageHandler(
        automation, script_data, redpacket_config
    )
    
    # 設置消息事件處理
    for client, user_id, config in connected:
        @client.on(events.NewMessage)
        async def handler(event, uid=user_id, cli=client):
            await message_handler.handle_new_message(cli, uid, event)
    
    # 自動建群
    test_group = None
    if AUTO_CREATE_GROUP:
        print("\n🏠 正在創建業務群組...")
        
        creator_user_id = connected[0][1]
        
        test_group = await automation.create_and_start_group(
            creator_user_id=creator_user_id,
            group_name=GROUP_NAME if GROUP_NAME else None,
            script_id=SCRIPT_ID
        )
        
        if test_group:
            print()
            print("=" * 50)
            print(f"🎉 業務群組創建成功！")
            print(f"   名稱: {test_group.group_name}")
            print(f"   ID: {test_group.group_id}")
            print(f"   邀請鏈接: {test_group.invite_link}")
            print(f"   當前階段: {test_group.current_stage.value}")
            print("=" * 50)
            print()
            
            # 啟動自動邀請
            if AUTO_INVITE:
                await automation.start_auto_invitation(test_group.group_id)
    
    # 狀態顯示
    print()
    print("🚀 系統已啟動！")
    print(f"   📊 {len(connected)} 個帳號在線")
    print(f"   🧧 自動搶紅包: {'✅' if AUTO_GRAB else '❌'}")
    print(f"   📤 自動發紅包: {'✅' if AUTO_SEND else '❌'}")
    print(f"   💬 劇本驅動對話: {'✅' if script_data else '❌'}")
    print(f"   👥 自動邀請用戶: {'✅' if AUTO_INVITE else '❌'}")
    if test_group:
        print(f"   🏠 活躍群組: {test_group.group_name}")
    print()
    print("按 Ctrl+C 停止")
    print()
    
    # 保持運行
    try:
        while True:
            await asyncio.sleep(60)
            
            # 輸出狀態
            status = automation.get_system_status()
            
            for group_info in status.get("active_groups", []):
                if group_info:
                    logger.info(
                        f"📊 群組 {group_info['group_name']}: "
                        f"階段={group_info['stage']}, "
                        f"AI={group_info['ai_count']}, "
                        f"用戶={group_info['user_count']}, "
                        f"消息={group_info['total_messages']}"
                    )
    
    except KeyboardInterrupt:
        print("\n正在停止...")
    finally:
        # 斷開所有客戶端
        for client, _, _ in connected:
            try:
                await client.disconnect()
            except:
                pass
        
        print("✅ 已停止")


if __name__ == "__main__":
    asyncio.run(main())
