#!/usr/bin/env python3
"""
🚀 Worker 節點部署腳本

功能：
1. 檢查環境依賴
2. 創建必要目錄
3. 生成配置文件
4. 測試 API 連通性
5. 驗證帳號配置

使用方法:
    python deploy_worker.py
"""

import os
import sys
import subprocess
import json
from pathlib import Path
from datetime import datetime

# 顏色輸出
class Colors:
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    BLUE = '\033[94m'
    RESET = '\033[0m'
    BOLD = '\033[1m'

def print_header(text):
    print(f"\n{Colors.BOLD}{Colors.BLUE}{'='*60}{Colors.RESET}")
    print(f"{Colors.BOLD}{Colors.BLUE}  {text}{Colors.RESET}")
    print(f"{Colors.BOLD}{Colors.BLUE}{'='*60}{Colors.RESET}\n")

def print_success(text):
    print(f"{Colors.GREEN}✅ {text}{Colors.RESET}")

def print_warning(text):
    print(f"{Colors.YELLOW}⚠️  {text}{Colors.RESET}")

def print_error(text):
    print(f"{Colors.RED}❌ {text}{Colors.RESET}")

def print_info(text):
    print(f"{Colors.BLUE}ℹ️  {text}{Colors.RESET}")


# ==================== 檢查依賴 ====================

def check_python_version():
    """檢查 Python 版本"""
    print_info("檢查 Python 版本...")
    
    version = sys.version_info
    if version.major >= 3 and version.minor >= 9:
        print_success(f"Python 版本: {version.major}.{version.minor}.{version.micro}")
        return True
    else:
        print_error(f"需要 Python 3.9+，當前: {version.major}.{version.minor}")
        return False


def check_pip_packages():
    """檢查必要的 pip 包"""
    print_info("檢查依賴包...")
    
    required = [
        "telethon",
        "httpx",
        "openpyxl",
        "pyyaml",
        "fastapi",
        "uvicorn"
    ]
    
    missing = []
    
    for package in required:
        try:
            __import__(package.replace("-", "_"))
            print_success(f"  {package} ✓")
        except ImportError:
            print_warning(f"  {package} - 未安裝")
            missing.append(package)
    
    return missing


def install_packages(packages):
    """安裝缺失的包"""
    if not packages:
        return True
    
    print_info(f"安裝缺失的包: {', '.join(packages)}")
    
    try:
        subprocess.check_call([
            sys.executable, "-m", "pip", "install", 
            "-r", "requirements.txt",
            "-q"
        ])
        print_success("依賴安裝完成")
        return True
    except subprocess.CalledProcessError as e:
        print_error(f"安裝失敗: {e}")
        return False


# ==================== 創建目錄 ====================

def create_directories():
    """創建必要的目錄"""
    print_info("創建目錄結構...")
    
    dirs = [
        "sessions",
        "logs",
        "scripts",
        "config",
        "data"
    ]
    
    for d in dirs:
        path = Path(d)
        path.mkdir(parents=True, exist_ok=True)
        print_success(f"  {d}/")
    
    return True


# ==================== 生成配置 ====================

def create_env_file():
    """創建 .env 配置文件"""
    env_file = Path(".env.worker")
    
    if env_file.exists():
        print_warning(".env.worker 已存在，跳過")
        return True
    
    print_info("創建環境配置文件...")
    
    env_content = """# Worker 節點配置文件
# 生成時間: {timestamp}

# ========== 紅包 API 配置 ==========
REDPACKET_API_URL=http://api.usdt2026.cc
REDPACKET_API_KEY=test-key-2024

# ========== 遊戲策略 ==========
# conservative / balanced / aggressive / smart
GAME_STRATEGY=smart

# ========== 自動化設置 ==========
AUTO_CREATE_GROUP=true
AUTO_GRAB=true
AUTO_SEND=false
AUTO_CHAT=true

# ========== LLM 配置（可選）==========
LLM_ENABLED=false
# OPENAI_API_KEY=sk-xxx
# LLM_PROVIDER=openai
# LLM_MODEL=gpt-3.5-turbo

# ========== 目錄配置 ==========
SESSIONS_DIR=./sessions
SCRIPTS_DIR=./scripts
LOG_DIR=./logs

# ========== 日誌配置 ==========
LOG_LEVEL=INFO
""".format(timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    
    env_file.write_text(env_content, encoding="utf-8")
    print_success(".env.worker 已創建")
    
    return True


def create_sample_accounts_excel():
    """創建示例帳號配置 Excel"""
    excel_file = Path("sessions/accounts_sample.xlsx")
    
    if excel_file.exists():
        print_warning("accounts_sample.xlsx 已存在，跳過")
        return True
    
    print_info("創建示例帳號配置...")
    
    try:
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AI帳號配置"
        
        # 表頭
        headers = [
            "phone", "api_id", "api_hash", "name", 
            "role", "strategy", "status", "notes"
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 示例數據 (使用文檔中的 AI 帳號)
        sample_data = [
            ["639277358115", "YOUR_API_ID_1", "YOUR_API_HASH_1", "AI-1", "xiaoqi", "smart", "active", "測試帳號1"],
            ["639543603735", "YOUR_API_ID_2", "YOUR_API_HASH_2", "AI-2", "mimi", "smart", "active", "測試帳號2"],
            ["639952948692", "YOUR_API_ID_3", "YOUR_API_HASH_3", "AI-3", "haoge", "smart", "active", "測試帳號3"],
            ["639454959591", "YOUR_API_ID_4", "YOUR_API_HASH_4", "AI-4", "xiaoyu", "smart", "active", "測試帳號4"],
            ["639542360349", "YOUR_API_ID_5", "YOUR_API_HASH_5", "AI-5", "aqiang", "smart", "active", "測試帳號5"],
            ["639950375245", "YOUR_API_ID_6", "YOUR_API_HASH_6", "AI-6", "laozhang", "smart", "active", "測試帳號6"],
        ]
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 調整列寬
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 20
        
        wb.save(excel_file)
        print_success("accounts_sample.xlsx 已創建")
        print_warning("請編輯此文件，填入真實的 API_ID 和 API_HASH")
        
        return True
        
    except Exception as e:
        print_error(f"創建 Excel 失敗: {e}")
        return False


# ==================== API 測試 ====================

def test_redpacket_api():
    """測試紅包 API 連通性"""
    print_info("測試紅包 API 連通性...")
    
    try:
        import httpx
        
        response = httpx.get(
            "http://api.usdt2026.cc/api/v2/ai/status",
            timeout=10
        )
        
        if response.status_code == 200:
            print_success("紅包 API 連接正常")
            return True
        else:
            print_warning(f"API 響應異常: {response.status_code}")
            return False
            
    except Exception as e:
        print_error(f"API 連接失敗: {e}")
        return False


def test_ai_account_balance():
    """測試 AI 帳號餘額"""
    print_info("測試 AI 帳號餘額查詢...")
    
    try:
        import httpx
        
        # 測試第一個 AI 帳號
        ai_id = 639277358115
        
        response = httpx.get(
            "http://api.usdt2026.cc/api/v2/ai/wallet/balance",
            headers={
                "Authorization": "Bearer test-key-2024",
                "X-Telegram-User-Id": str(ai_id)
            },
            timeout=10
        )
        
        if response.status_code == 200:
            data = response.json()
            balance = data.get("data", {}).get("balances", {}).get("usdt", 0)
            print_success(f"AI-1 (ID: {ai_id}) 餘額: {balance} USDT")
            return True
        else:
            print_warning(f"查詢失敗: {response.status_code}")
            return False
            
    except Exception as e:
        print_error(f"查詢失敗: {e}")
        return False


# ==================== 主程序 ====================

def main():
    print_header("🚀 Worker 節點部署工具")
    
    # 切換到腳本目錄
    script_dir = Path(__file__).parent
    os.chdir(script_dir)
    
    print(f"工作目錄: {os.getcwd()}\n")
    
    # 步驟 1: 檢查 Python 版本
    print_header("步驟 1: 環境檢查")
    if not check_python_version():
        sys.exit(1)
    
    # 步驟 2: 檢查依賴
    missing = check_pip_packages()
    if missing:
        print()
        if input("是否安裝缺失的依賴？(y/n): ").lower() == 'y':
            if not install_packages(missing):
                print_error("依賴安裝失敗，請手動執行: pip install -r requirements.txt")
                sys.exit(1)
    
    # 步驟 3: 創建目錄
    print_header("步驟 2: 創建目錄結構")
    create_directories()
    
    # 步驟 4: 生成配置
    print_header("步驟 3: 生成配置文件")
    create_env_file()
    create_sample_accounts_excel()
    
    # 步驟 5: API 測試
    print_header("步驟 4: API 連通性測試")
    api_ok = test_redpacket_api()
    if api_ok:
        test_ai_account_balance()
    
    # 完成
    print_header("✅ 部署準備完成")
    
    print("""
下一步操作：

1️⃣  編輯帳號配置
    打開 sessions/accounts_sample.xlsx
    填入每個帳號的真實 API_ID 和 API_HASH
    重命名為 accounts.xlsx

2️⃣  放入 Session 文件
    將 .session 文件複製到 sessions/ 目錄
    文件名應包含電話號碼（如 639277358115.session）

3️⃣  配置環境變量（可選）
    編輯 .env.worker 文件
    根據需要調整設置

4️⃣  啟動系統
    python start_full_system.py

📌 AI 帳號列表（來自 API 文檔）：
   - AI-1: 639277358115
   - AI-2: 639543603735
   - AI-3: 639952948692
   - AI-4: 639454959591
   - AI-5: 639542360349
   - AI-6: 639950375245

📖 完整文檔：docs/完整系統功能說明.md
""")


if __name__ == "__main__":
    main()
