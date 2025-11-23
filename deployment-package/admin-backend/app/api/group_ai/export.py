"""
數據導出 API
支持導出劇本、賬號、分配方案、監控數據等為 Excel、PDF、CSV 格式
"""
import logging
import io
import csv
from typing import List, Dict, Any, Optional
from datetime import datetime
from fastapi import APIRouter, HTTPException, Query, Depends
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session

from app.db import get_db
from app.models.group_ai import (
    GroupAIScript,
    GroupAIAccount,
    GroupAIRoleAssignmentScheme,
)
from app.api.deps import get_optional_user
from app.models.user import User

logger = logging.getLogger(__name__)

router = APIRouter(tags=["Data Export"])

# 嘗試導入 Excel 和 PDF 庫
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl 未安裝，Excel 導出功能不可用")

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger.warning("reportlab 未安裝，PDF 導出功能不可用")


def export_to_csv(data: List[Dict[str, Any]], filename: str) -> StreamingResponse:
    """導出數據為 CSV 格式"""
    output = io.StringIO()
    
    if not data:
        # 空數據，只寫標題
        writer = csv.writer(output)
        writer.writerow(["無數據"])
    else:
        # 獲取所有可能的字段
        fieldnames = set()
        for row in data:
            fieldnames.update(row.keys())
        fieldnames = sorted(fieldnames)
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        for row in data:
            # 處理嵌套字典和列表
            clean_row = {}
            for key, value in row.items():
                if isinstance(value, (dict, list)):
                    clean_row[key] = str(value)
                elif value is None:
                    clean_row[key] = ""
                else:
                    clean_row[key] = value
            writer.writerow(clean_row)
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}.csv"'
        }
    )


def export_to_excel(data: List[Dict[str, Any]], filename: str, sheet_name: str = "數據") -> StreamingResponse:
    """導出數據為 Excel 格式"""
    if not EXCEL_AVAILABLE:
        raise HTTPException(
            status_code=503,
            detail="Excel 導出功能不可用，請安裝 openpyxl: pip install openpyxl"
        )
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    if not data:
        ws.append(["無數據"])
    else:
        # 獲取所有字段
        fieldnames = set()
        for row in data:
            fieldnames.update(row.keys())
        fieldnames = sorted(fieldnames)
        
        # 寫入標題行
        header_row = [field for field in fieldnames]
        ws.append(header_row)
        
        # 設置標題樣式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 寫入數據行
        for row in data:
            row_data = []
            for field in fieldnames:
                value = row.get(field, "")
                if isinstance(value, (dict, list)):
                    value = str(value)
                elif value is None:
                    value = ""
                row_data.append(value)
            ws.append(row_data)
        
        # 自動調整列寬
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}.xlsx"'
        }
    )


def export_to_pdf(data: List[Dict[str, Any]], filename: str, title: str = "數據導出") -> StreamingResponse:
    """導出數據為 PDF 格式"""
    if not PDF_AVAILABLE:
        raise HTTPException(
            status_code=503,
            detail="PDF 導出功能不可用，請安裝 reportlab: pip install reportlab"
        )
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    
    # 標題
    title_para = Paragraph(title, styles['Title'])
    elements.append(title_para)
    elements.append(Spacer(1, 0.2*inch))
    
    # 導出時間
    time_para = Paragraph(f"導出時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])
    elements.append(time_para)
    elements.append(Spacer(1, 0.2*inch))
    
    if not data:
        no_data_para = Paragraph("無數據", styles['Normal'])
        elements.append(no_data_para)
    else:
        # 獲取所有字段
        fieldnames = set()
        for row in data:
            fieldnames.update(row.keys())
        fieldnames = sorted(fieldnames)
        
        # 準備表格數據
        table_data = [fieldnames]  # 標題行
        
        for row in data:
            row_data = []
            for field in fieldnames:
                value = row.get(field, "")
                if isinstance(value, (dict, list)):
                    value = str(value)
                elif value is None:
                    value = ""
                else:
                    value = str(value)
                row_data.append(value)
            table_data.append(row_data)
        
        # 創建表格
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
    
    # 構建 PDF
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        io.BytesIO(buffer.read()),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}.pdf"'
        }
    )


@router.get("/scripts")
async def export_scripts(
    format: str = Query(..., description="導出格式: csv, excel, pdf"),
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(get_optional_user)
):
    """導出劇本列表"""
    try:
        scripts = db.query(GroupAIScript).order_by(GroupAIScript.created_at.desc()).all()
        
        data = []
        for script in scripts:
            data.append({
                "劇本ID": script.script_id,
                "名稱": script.name,
                "版本": script.version,
                "描述": script.description or "",
                "狀態": script.status,
                "創建者": script.created_by or "",
                "創建時間": script.created_at.strftime("%Y-%m-%d %H:%M:%S") if script.created_at else "",
                "更新時間": script.updated_at.strftime("%Y-%m-%d %H:%M:%S") if script.updated_at else "",
            })
        
        filename = f"劇本列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        if format.lower() == "csv":
            return export_to_csv(data, filename)
        elif format.lower() in ["excel", "xlsx"]:
            return export_to_excel(data, filename, "劇本列表")
        elif format.lower() == "pdf":
            return export_to_pdf(data, filename, "劇本列表")
        else:
            raise HTTPException(status_code=400, detail=f"不支持的格式: {format}")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"導出劇本列表失敗: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"導出失敗: {str(e)}")


@router.get("/accounts")
async def export_accounts(
    format: str = Query(..., description="導出格式: csv, excel, pdf"),
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(get_optional_user)
):
    """導出賬號列表"""
    try:
        accounts = db.query(GroupAIAccount).order_by(GroupAIAccount.created_at.desc()).all()
        
        data = []
        for account in accounts:
            data.append({
                "賬號ID": account.account_id,
                "顯示名稱": account.display_name or account.account_id,
                "用戶名": account.username or "",
                "手機號": account.phone_number or "",
                "劇本ID": account.script_id,
                "服務器ID": account.server_id or "",
                "狀態": "在線" if account.active else "離線",
                "群組數量": len(account.group_ids) if account.group_ids else 0,
                "回復率": account.reply_rate,
                "紅包啟用": "是" if account.redpacket_enabled else "否",
                "創建時間": account.created_at.strftime("%Y-%m-%d %H:%M:%S") if account.created_at else "",
            })
        
        filename = f"賬號列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        if format.lower() == "csv":
            return export_to_csv(data, filename)
        elif format.lower() in ["excel", "xlsx"]:
            return export_to_excel(data, filename, "賬號列表")
        elif format.lower() == "pdf":
            return export_to_pdf(data, filename, "賬號列表")
        else:
            raise HTTPException(status_code=400, detail=f"不支持的格式: {format}")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"導出賬號列表失敗: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"導出失敗: {str(e)}")


@router.get("/schemes")
async def export_schemes(
    format: str = Query(..., description="導出格式: csv, excel, pdf"),
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(get_optional_user)
):
    """導出分配方案列表"""
    try:
        schemes = db.query(GroupAIRoleAssignmentScheme).order_by(
            GroupAIRoleAssignmentScheme.created_at.desc()
        ).all()
        
        data = []
        for scheme in schemes:
            # 獲取劇本名稱
            script = db.query(GroupAIScript).filter(
                GroupAIScript.script_id == scheme.script_id
            ).first()
            script_name = script.name if script else scheme.script_id
            
            data.append({
                "方案ID": scheme.id,
                "方案名稱": scheme.name,
                "描述": scheme.description or "",
                "劇本ID": scheme.script_id,
                "劇本名稱": script_name,
                "分配模式": scheme.mode,
                "賬號數量": len(scheme.account_ids) if scheme.account_ids else 0,
                "分配數量": len(scheme.assignments) if scheme.assignments else 0,
                "創建者": scheme.created_by or "",
                "創建時間": scheme.created_at.strftime("%Y-%m-%d %H:%M:%S") if scheme.created_at else "",
                "更新時間": scheme.updated_at.strftime("%Y-%m-%d %H:%M:%S") if scheme.updated_at else "",
            })
        
        filename = f"分配方案列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        if format.lower() == "csv":
            return export_to_csv(data, filename)
        elif format.lower() in ["excel", "xlsx"]:
            return export_to_excel(data, filename, "分配方案列表")
        elif format.lower() == "pdf":
            return export_to_pdf(data, filename, "分配方案列表")
        else:
            raise HTTPException(status_code=400, detail=f"不支持的格式: {format}")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"導出分配方案列表失敗: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"導出失敗: {str(e)}")


@router.get("/schemes/{scheme_id}/details")
async def export_scheme_details(
    scheme_id: str,
    format: str = Query(..., description="導出格式: csv, excel, pdf"),
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(get_optional_user)
):
    """導出分配方案詳情（包含所有分配項）"""
    try:
        scheme = db.query(GroupAIRoleAssignmentScheme).filter(
            GroupAIRoleAssignmentScheme.id == scheme_id
        ).first()
        
        if not scheme:
            raise HTTPException(status_code=404, detail="分配方案不存在")
        
        # 獲取劇本名稱
        script = db.query(GroupAIScript).filter(
            GroupAIScript.script_id == scheme.script_id
        ).first()
        script_name = script.name if script else scheme.script_id
        
        data = []
        for assignment in (scheme.assignments or []):
            data.append({
                "方案名稱": scheme.name,
                "劇本ID": scheme.script_id,
                "劇本名稱": script_name,
                "角色ID": assignment.get("role_id", ""),
                "角色名稱": assignment.get("role_name", ""),
                "賬號ID": assignment.get("account_id", ""),
                "權重": assignment.get("weight", ""),
            })
        
        filename = f"分配方案詳情_{scheme.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        if format.lower() == "csv":
            return export_to_csv(data, filename)
        elif format.lower() in ["excel", "xlsx"]:
            return export_to_excel(data, filename, "分配詳情")
        elif format.lower() == "pdf":
            return export_to_pdf(data, filename, f"分配方案詳情: {scheme.name}")
        else:
            raise HTTPException(status_code=400, detail=f"不支持的格式: {format}")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"導出分配方案詳情失敗: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"導出失敗: {str(e)}")

