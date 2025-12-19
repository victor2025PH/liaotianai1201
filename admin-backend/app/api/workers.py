"""
Workers API - 分布式节点管理系统
用于管理本地电脑和远程服务器节点
"""

import logging
import json
import time
import base64
import asyncio
from typing import Dict, List, Optional, Any
from datetime import datetime, timedelta
from pathlib import Path
from fastapi import APIRouter, Depends, HTTPException, status, Body, UploadFile, File
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from app.api.deps import get_db_session, get_current_active_user
from app.models.user import User
from app.core.config import get_settings

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/workers", tags=["workers"])

# 内存存储（当 Redis 不可用时使用）
_workers_memory_store: Dict[str, Dict[str, Any]] = {}
_worker_commands: Dict[str, List[Dict[str, Any]]] = {}  # 存储待执行的命令
_worker_responses: Dict[str, Dict[str, Any]] = {}  # 存储 Worker 节点的响应结果

# Redis 客户端（如果可用）
_redis_client = None
_redis_pubsub = None
try:
    import redis
    settings = get_settings()
    if settings.redis_url:
        _redis_client = redis.from_url(settings.redis_url, decode_responses=True)
        _redis_client.ping()
        _redis_pubsub = _redis_client.pubsub()
        logger.info("Workers API: Redis 已启用")
    else:
        logger.info("Workers API: Redis 未配置，使用内存存储")
except Exception as e:
    logger.warning(f"Workers API: Redis 不可用，使用内存存储: {e}")
    _redis_client = None
    _redis_pubsub = None


# ============ 数据模型 ============

class WorkerHeartbeatRequest(BaseModel):
    """Worker 节点心跳请求"""
    node_id: str = Field(..., description="节点ID，如 computer_001, computer_002")
    status: str = Field(default="online", description="节点状态: online, offline")
    account_count: int = Field(default=0, description="账号数量")
    accounts: List[Dict[str, Any]] = Field(default_factory=list, description="账号列表")
    metadata: Optional[Dict[str, Any]] = Field(default=None, description="额外元数据")
    # 新增：Worker 节点可以返回命令执行结果
    command_responses: Optional[Dict[str, Any]] = Field(default=None, description="命令执行结果")


class WorkerCommandRequest(BaseModel):
    """向 Worker 节点发送命令请求"""
    action: str = Field(..., description="命令动作，如 start_auto_chat, stop_auto_chat, set_config, create_group")
    params: Dict[str, Any] = Field(default_factory=dict, description="命令参数")


class WorkerBroadcastRequest(BaseModel):
    """广播命令到所有节点请求"""
    action: str = Field(..., description="命令动作")
    params: Dict[str, Any] = Field(default_factory=dict, description="命令参数")


class WorkerStatus(BaseModel):
    """Worker 节点状态"""
    node_id: str
    status: str
    account_count: int
    last_heartbeat: str
    accounts: List[Dict[str, Any]] = Field(default_factory=list)
    metadata: Optional[Dict[str, Any]] = None


class WorkersListResponse(BaseModel):
    """Workers 列表响应"""
    workers: Dict[str, WorkerStatus]


class SessionFileInfo(BaseModel):
    """Session 文件信息"""
    filename: str
    size: int
    modified_time: Optional[str] = None
    path: Optional[str] = None


class SessionListResponse(BaseModel):
    """Session 列表响应"""
    success: bool
    node_id: str
    sessions: List[SessionFileInfo]
    total: int
    message: str


class SessionUploadResponse(BaseModel):
    """Session 上传响应"""
    success: bool
    node_id: str
    filename: str
    message: str


class SessionDownloadResponse(BaseModel):
    """Session 下载响应"""
    success: bool
    node_id: str
    filename: str
    file_content: str  # base64 编码的文件内容
    file_size: int
    message: str


class SessionDeleteResponse(BaseModel):
    """Session 删除响应"""
    success: bool
    node_id: str
    filename: str
    message: str


class BatchOperationRequest(BaseModel):
    """批量操作请求"""
    filenames: List[str] = Field(..., description="文件名列表")
    operation: str = Field(..., description="操作类型: upload, download, delete")


class BatchOperationResponse(BaseModel):
    """批量操作响应"""
    success: bool
    node_id: str
    operation: str
    total: int
    successful: int
    failed: int
    results: List[Dict[str, Any]]
    message: str


# ============ 辅助函数 ============

def _get_worker_key(node_id: str) -> str:
    """获取 Worker 在 Redis 中的键"""
    return f"worker:node:{node_id}"


def _get_workers_set_key() -> str:
    """获取所有 Worker 节点集合的键"""
    return "worker:nodes:all"


def _get_commands_key(node_id: str) -> str:
    """获取节点命令队列的键"""
    return f"worker:commands:{node_id}"


def _get_response_key(node_id: str, command_id: str) -> str:
    """获取命令响应存储的键"""
    return f"worker:response:{node_id}:{command_id}"


def _save_worker_status(node_id: str, data: Dict[str, Any]) -> None:
    """保存 Worker 节点状态"""
    data["last_heartbeat"] = datetime.now().isoformat()
    
    if _redis_client:
        try:
            key = _get_worker_key(node_id)
            _redis_client.setex(key, 120, json.dumps(data))
            _redis_client.sadd(_get_workers_set_key(), node_id)
            _redis_client.expire(_get_workers_set_key(), 120)
        except Exception as e:
            logger.error(f"保存 Worker 状态到 Redis 失败: {e}")
            _workers_memory_store[node_id] = data
    else:
        _workers_memory_store[node_id] = data


def _get_worker_status(node_id: str) -> Optional[Dict[str, Any]]:
    """获取 Worker 节点状态"""
    if _redis_client:
        try:
            key = _get_worker_key(node_id)
            data = _redis_client.get(key)
            if data:
                return json.loads(data)
        except Exception as e:
            logger.error(f"从 Redis 获取 Worker 状态失败: {e}")
            return _workers_memory_store.get(node_id)
    else:
        return _workers_memory_store.get(node_id)


def _get_all_workers() -> Dict[str, Dict[str, Any]]:
    """获取所有 Worker 节点状态，并检查心跳超时"""
    workers = {}
    HEARTBEAT_TIMEOUT_SECONDS = 90  # 心跳超时时间：90秒
    
    if _redis_client:
        try:
            node_ids = _redis_client.smembers(_get_workers_set_key())
            for node_id in node_ids:
                status_data = _get_worker_status(node_id)
                if status_data:
                    # 检查心跳是否超时
                    last_heartbeat_str = status_data.get("last_heartbeat")
                    original_status = status_data.get("status", "offline")  # 保存原始状态
                    
                    if last_heartbeat_str:
                        try:
                            last_heartbeat = datetime.fromisoformat(last_heartbeat_str.replace('Z', '+00:00'))
                            # 处理时区
                            if last_heartbeat.tzinfo is None:
                                last_heartbeat = last_heartbeat.replace(tzinfo=datetime.now().astimezone().tzinfo)
                            
                            time_since_heartbeat = (datetime.now().astimezone() - last_heartbeat).total_seconds()
                            
                            # 如果心跳超时，标记为离线（覆盖原始状态）
                            if time_since_heartbeat > HEARTBEAT_TIMEOUT_SECONDS:
                                status_data["status"] = "offline"
                                logger.debug(f"节点 {node_id} 心跳超时 ({time_since_heartbeat:.0f}秒)，标记为离线")
                            else:
                                # 心跳未超时，使用原始状态（可能是 "online"）
                                status_data["status"] = original_status if original_status else "online"
                        except (ValueError, TypeError) as e:
                            logger.warning(f"解析节点 {node_id} 心跳时间失败: {e}")
                            status_data["status"] = "offline"
                    else:
                        # 没有心跳时间，标记为离线
                        status_data["status"] = "offline"
                    
                    workers[node_id] = status_data
        except Exception as e:
            logger.error(f"从 Redis 获取所有 Workers 失败: {e}")
            workers = _workers_memory_store.copy()
            # 对内存存储也进行超时检测
            _check_heartbeat_timeout(workers, HEARTBEAT_TIMEOUT_SECONDS)
    else:
        workers = _workers_memory_store.copy()
        # 对内存存储也进行超时检测
        _check_heartbeat_timeout(workers, HEARTBEAT_TIMEOUT_SECONDS)
    
    return workers


def _check_heartbeat_timeout(workers: Dict[str, Dict[str, Any]], timeout_seconds: int) -> None:
    """检查心跳超时并更新状态（辅助函数）"""
    for node_id, status_data in workers.items():
        last_heartbeat_str = status_data.get("last_heartbeat")
        if last_heartbeat_str:
            try:
                last_heartbeat = datetime.fromisoformat(last_heartbeat_str.replace('Z', '+00:00'))
                if last_heartbeat.tzinfo is None:
                    last_heartbeat = last_heartbeat.replace(tzinfo=datetime.now().astimezone().tzinfo)
                
                time_since_heartbeat = (datetime.now().astimezone() - last_heartbeat).total_seconds()
                
                if time_since_heartbeat > timeout_seconds:
                    status_data["status"] = "offline"
                    logger.debug(f"节点 {node_id} 心跳超时 ({time_since_heartbeat:.0f}秒)，标记为离线")
            except (ValueError, TypeError) as e:
                logger.warning(f"解析节点 {node_id} 心跳时间失败: {e}")
                status_data["status"] = "offline"
        else:
            status_data["status"] = "offline"


def _add_command(node_id: str, command: Dict[str, Any]) -> None:
    """添加命令到节点命令队列"""
    if _redis_client:
        try:
            key = _get_commands_key(node_id)
            _redis_client.lpush(key, json.dumps(command))
            _redis_client.expire(key, 300)  # 命令队列TTL: 5分钟
        except Exception as e:
            logger.error(f"添加命令到 Redis 失败: {e}")
            # 降级到内存存储
            if node_id not in _worker_commands:
                _worker_commands[node_id] = []
            _worker_commands[node_id].append(command)
    else:
        # 使用内存存储
        if node_id not in _worker_commands:
            _worker_commands[node_id] = []
        _worker_commands[node_id].append(command)


def _get_commands(node_id: str) -> List[Dict[str, Any]]:
    """获取节点的待执行命令"""
    if _redis_client:
        try:
            key = _get_commands_key(node_id)
            commands = _redis_client.lrange(key, 0, -1)
            return [json.loads(cmd) for cmd in commands]
        except Exception as e:
            logger.error(f"从 Redis 获取命令失败: {e}")
            return _worker_commands.get(node_id, [])
    else:
        return _worker_commands.get(node_id, [])


def _clear_commands(node_id: str) -> None:
    """清除节点的命令队列"""
    if _redis_client:
        try:
            key = _get_commands_key(node_id)
            _redis_client.delete(key)
        except Exception as e:
            logger.error(f"清除 Redis 命令队列失败: {e}")
            if node_id in _worker_commands:
                del _worker_commands[node_id]
    else:
        if node_id in _worker_commands:
            del _worker_commands[node_id]


def _save_response(node_id: str, command_id: str, response: Dict[str, Any]) -> None:
    """保存 Worker 节点的响应结果"""
    response["timestamp"] = datetime.now().isoformat()
    
    if _redis_client:
        try:
            key = _get_response_key(node_id, command_id)
            _redis_client.setex(key, 60, json.dumps(response))  # TTL: 60秒
        except Exception as e:
            logger.error(f"保存响应到 Redis 失败: {e}")
            if node_id not in _worker_responses:
                _worker_responses[node_id] = {}
            _worker_responses[node_id][command_id] = response
    else:
        if node_id not in _worker_responses:
            _worker_responses[node_id] = {}
        _worker_responses[node_id][command_id] = response


def _get_response(node_id: str, command_id: str) -> Optional[Dict[str, Any]]:
    """获取 Worker 节点的响应结果"""
    if _redis_client:
        try:
            key = _get_response_key(node_id, command_id)
            data = _redis_client.get(key)
            if data:
                return json.loads(data)
        except Exception as e:
            logger.error(f"从 Redis 获取响应失败: {e}")
            return _worker_responses.get(node_id, {}).get(command_id)
    else:
        return _worker_responses.get(node_id, {}).get(command_id)


def _wait_for_response(node_id: str, command_id: str, timeout: int = 30) -> Optional[Dict[str, Any]]:
    """等待 Worker 节点的响应（轮询方式）"""
    start_time = time.time()
    while time.time() - start_time < timeout:
        response = _get_response(node_id, command_id)
        if response:
            return response
        time.sleep(0.5)  # 每 0.5 秒检查一次
    return None


# ============ API 端点 ============

@router.post("/heartbeat", status_code=status.HTTP_200_OK)
async def worker_heartbeat(
    request: WorkerHeartbeatRequest,
    db: Session = Depends(get_db_session)
):
    """
    Worker 节点心跳端点
    节点应每30秒调用一次此端点来报告状态
    
    新增功能：Worker 节点可以通过 command_responses 字段返回命令执行结果
    """
    try:
        worker_data = {
            "node_id": request.node_id,
            "status": request.status or "online",  # 如果未提供状态，默认为 online
            "account_count": request.account_count,
            "accounts": request.accounts,
            "metadata": request.metadata or {},
            "last_heartbeat": datetime.now().isoformat()
        }
        
        _save_worker_status(request.node_id, worker_data)
        
        # 处理命令执行结果
        if request.command_responses:
            for command_id, response_data in request.command_responses.items():
                _save_response(request.node_id, command_id, response_data)
                logger.info(f"收到节点 {request.node_id} 的命令响应: {command_id}")
        
        # 同步賬號信息到數據庫
        if request.accounts:
            try:
                from app.api.group_ai.remote_account_sync import sync_accounts_from_worker
                sync_result = sync_accounts_from_worker(
                    node_id=request.node_id,
                    accounts=request.accounts,
                    db=db
                )
                logger.info(f"從節點 {request.node_id} 同步了 {sync_result['synced_count']} 個賬號")
            except Exception as sync_error:
                logger.error(f"同步賬號信息失敗: {sync_error}", exc_info=True)
        
        # 检查是否有待执行的命令
        commands = _get_commands(request.node_id)
        
        logger.info(f"Worker {request.node_id} 心跳: {request.account_count} 账号, {len(commands)} 待执行命令")
        
        return {
            "success": True,
            "node_id": request.node_id,
            "pending_commands": commands,
            "message": "心跳已接收"
        }
    except Exception as e:
        logger.error(f"处理 Worker 心跳失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"处理心跳失败: {str(e)}"
        )


@router.get("/", response_model=WorkersListResponse)
async def list_workers(
    current_user: Optional[User] = Depends(get_current_active_user),
    db: Session = Depends(get_db_session)
):
    """
    获取所有 Worker 节点状态列表
    注意：如果禁用认证（DISABLE_AUTH=true），则允许匿名访问
    """
    # 如果禁用认证，current_user 可能为 None，这是允许的
    """
    获取所有 Worker 节点状态列表
    """
    try:
        workers_data = _get_all_workers()
        
        # 转换为响应格式，并确保状态正确（基于心跳超时）
        workers = {}
        for node_id, data in workers_data.items():
            # _get_all_workers() 已经检查了心跳超时，直接使用其返回的状态
            status = data.get("status", "offline")
            
            workers[node_id] = WorkerStatus(
                node_id=node_id,
                status=status,  # 使用经过心跳超时检测的状态
                account_count=data.get("account_count", 0),
                last_heartbeat=data.get("last_heartbeat", datetime.now().isoformat()),
                accounts=data.get("accounts", []),
                metadata=data.get("metadata")
            )
        
        return WorkersListResponse(workers=workers)
    except Exception as e:
        logger.error(f"获取 Workers 列表失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取 Workers 列表失败: {str(e)}"
        )


@router.get("/{node_id}/commands", status_code=status.HTTP_200_OK)
async def get_worker_commands(
    node_id: str,
    current_user: Optional[User] = Depends(get_current_active_user),
    db: Session = Depends(get_db_session)
):
    """
    获取节点的待执行命令（Worker 节点调用）
    """
    try:
        commands = _get_commands(node_id)
        return {
            "success": True,
            "node_id": node_id,
            "commands": commands,
            "count": len(commands)
        }
    except Exception as e:
        logger.error(f"获取命令失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取命令失败: {str(e)}"
        )


@router.post("/{node_id}/commands", status_code=status.HTTP_200_OK)
async def send_worker_command(
    node_id: str,
    request: WorkerCommandRequest,
    current_user: Optional[User] = Depends(get_current_active_user),
    db: Session = Depends(get_db_session)
):
    """
    向特定 Worker 节点发送命令
    """
    try:
        command = {
            "action": request.action,
            "params": request.params,
            "timestamp": datetime.now().isoformat(),
            "from": "master"
        }
        
        _add_command(node_id, command)
        
        logger.info(f"向节点 {node_id} 发送命令: {request.action}")
        
        return {
            "success": True,
            "node_id": node_id,
            "action": request.action,
            "message": f"命令已发送到节点 {node_id}"
        }
    except Exception as e:
        logger.error(f"发送命令到 Worker 失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"发送命令失败: {str(e)}"
        )


@router.post("/broadcast", status_code=status.HTTP_200_OK)
async def broadcast_command(
    request: WorkerBroadcastRequest,
    current_user: Optional[User] = Depends(get_current_active_user),
    db: Session = Depends(get_db_session)
):
    """
    广播命令到所有在线 Worker 节点
    """
    try:
        workers = _get_all_workers()
        online_workers = [node_id for node_id, data in workers.items() 
                         if data.get("status") == "online"]
        
        command = {
            "action": request.action,
            "params": request.params,
            "timestamp": datetime.now().isoformat(),
            "from": "master",
            "broadcast": True
        }
        
        for node_id in online_workers:
            _add_command(node_id, command)
        
        logger.info(f"广播命令 {request.action} 到 {len(online_workers)} 个节点")
        
        return {
            "success": True,
            "action": request.action,
            "nodes_count": len(online_workers),
            "nodes": online_workers,
            "message": f"命令已广播到 {len(online_workers)} 个节点"
        }
    except Exception as e:
        logger.error(f"广播命令失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"广播命令失败: {str(e)}"
        )


@router.delete("/{node_id}/commands", status_code=status.HTTP_200_OK)
async def clear_worker_commands(
    node_id: str,
    current_user: Optional[User] = Depends(get_current_active_user),
    db: Session = Depends(get_db_session)
):
    """
    清除节点的命令队列
    """
    try:
        _clear_commands(node_id)
        logger.info(f"已清除节点 {node_id} 的命令队列")
        
        return {
            "success": True,
            "node_id": node_id,
            "message": f"节点 {node_id} 的命令队列已清除"
        }
    except Exception as e:
        logger.error(f"清除命令队列失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"清除命令队列失败: {str(e)}"
        )


@router.delete("/{node_id}", status_code=status.HTTP_200_OK)
async def delete_worker(
    node_id: str,
    current_user: Optional[User] = Depends(get_current_active_user),
    db: Session = Depends(get_db_session)
):
    """
    刪除 Worker 節點
    """
    try:
        # 從 Redis 或內存中刪除
        if _redis_client:
            try:
                key = _get_worker_key(node_id)
                _redis_client.delete(key)
                _redis_client.srem(_get_workers_set_key(), node_id)
                # 清除命令隊列
                _redis_client.delete(_get_commands_key(node_id))
            except Exception as e:
                logger.error(f"從 Redis 刪除節點失敗: {e}")
        
        # 從內存存儲刪除
        if node_id in _workers_memory_store:
            del _workers_memory_store[node_id]
        if node_id in _worker_commands:
            del _worker_commands[node_id]
        
        logger.info(f"已刪除節點: {node_id}")
        
        return {
            "success": True,
            "node_id": node_id,
            "message": f"節點 {node_id} 已刪除"
        }
    except Exception as e:
        logger.error(f"刪除節點失敗: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"刪除節點失敗: {str(e)}"
        )


@router.get("/check/duplicates", status_code=status.HTTP_200_OK)
async def check_duplicate_accounts(
    current_user: Optional[User] = Depends(get_current_active_user),
    db: Session = Depends(get_db_session)
):
    """
    檢測重複帳號
    返回在多個節點上出現的帳號列表
    """
    try:
        workers_data = _get_all_workers()
        
        # 建立帳號到節點的映射
        account_nodes: Dict[str, List[str]] = {}
        
        for node_id, data in workers_data.items():
            accounts = data.get("accounts", [])
            for acc in accounts:
                acc_id = str(acc.get("account_id") or acc.get("user_id") or acc.get("phone", ""))
                if acc_id:
                    if acc_id not in account_nodes:
                        account_nodes[acc_id] = []
                    account_nodes[acc_id].append(node_id)
        
        # 找出重複帳號
        duplicates = []
        for acc_id, nodes in account_nodes.items():
            if len(nodes) > 1:
                duplicates.append({
                    "account_id": acc_id,
                    "nodes": nodes,
                    "count": len(nodes)
                })
        
        return {
            "success": True,
            "has_duplicates": len(duplicates) > 0,
            "duplicate_count": len(duplicates),
            "duplicates": duplicates,
            "total_accounts": len(account_nodes),
            "message": f"發現 {len(duplicates)} 個重複帳號" if duplicates else "無重複帳號"
        }
    except Exception as e:
        logger.error(f"檢測重複帳號失敗: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"檢測重複帳號失敗: {str(e)}"
        )


# ============ 部署包生成 API ============

class DeployPackageRequest(BaseModel):
    """部署包生成请求"""
    node_id: str = Field(..., description="节点ID")
    server_url: str = Field(..., description="服务器URL")
    api_key: str = Field(default="", description="API密钥（可选）")
    heartbeat_interval: int = Field(default=30, description="心跳间隔（秒）")
    telegram_api_id: str = Field(default="", description="Telegram API ID（可选，全局默认）")
    telegram_api_hash: str = Field(default="", description="Telegram API Hash（可选，全局默认）")


@router.post("/deploy-package", status_code=status.HTTP_200_OK)
async def generate_deploy_package(
    request: DeployPackageRequest,
    current_user: Optional[User] = Depends(get_current_active_user),
    db: Session = Depends(get_db_session)
):
    """
    生成 Worker 节点部署包
    
    返回包含所有必要脚本的部署包代码
    """
    try:
        # 生成 Windows 启动脚本（使用英文避免编码问题）
        windows_script = f"""@echo off
setlocal enabledelayedexpansion

REM ========================================
REM Worker Node Auto Deployment Script
REM Node ID: {request.node_id}
REM ========================================

REM Change to script directory
cd /d "%~dp0"

echo.
echo ========================================
echo Environment Check and Setup
echo ========================================
echo.

REM Step 1: Check Python
echo [1/4] Checking Python installation...
python --version >nul 2>&1
if errorlevel 1 (
    echo [ERROR] Python not found!
    echo.
    echo Please install Python 3.9 or higher:
    echo   1. Download from https://www.python.org/downloads/
    echo   2. During installation, check "Add Python to PATH"
    echo   3. Restart this script after installation
    echo.
    pause
    exit /b 1
)

for /f "tokens=2" %%i in ('python --version 2^>^&1') do set PYTHON_VERSION=%%i
echo [OK] Python found: !PYTHON_VERSION!

REM Step 2: Check and upgrade pip
echo.
echo [2/4] Checking pip...
python -m pip --version >nul 2>&1
if errorlevel 1 (
    echo [WARN] pip not found, installing...
    python -m ensurepip --upgrade
    if errorlevel 1 (
        echo [ERROR] Failed to install pip
        pause
        exit /b 1
    )
)

echo [OK] pip is available
echo [INFO] Upgrading pip to latest version...
python -m pip install --upgrade pip -i https://pypi.tuna.tsinghua.edu.cn/simple >nul 2>&1
if errorlevel 1 (
    python -m pip install --upgrade pip >nul 2>&1
)

REM Step 3: Check and install dependencies
echo.
echo [3/4] Checking Python dependencies...
set MISSING_DEPS=0

python -c "import telethon" >nul 2>&1
if errorlevel 1 (
    echo [MISSING] telethon
    set MISSING_DEPS=1
) else (
    echo [OK] telethon: installed
)

python -c "import requests" >nul 2>&1
if errorlevel 1 (
    echo [MISSING] requests
    set MISSING_DEPS=1
) else (
    echo [OK] requests: installed
)

python -c "import openpyxl" >nul 2>&1
if errorlevel 1 (
    echo [MISSING] openpyxl
    set MISSING_DEPS=1
) else (
    echo [OK] openpyxl: installed
)
if !MISSING_DEPS! equ 1 (
    echo.
    echo [INSTALL] Installing missing dependencies...
    echo [INFO] Using Tsinghua mirror source...
    python -m pip install telethon requests openpyxl -i https://pypi.tuna.tsinghua.edu.cn/simple
    if errorlevel 1 (
        echo [WARN] Mirror source failed, trying default source...
        python -m pip install telethon requests openpyxl --upgrade
        if errorlevel 1 (
            echo [ERROR] Failed to install dependencies
            pause
            exit /b 1
        )
    )
    echo.
    echo [VERIFY] Verifying installation...
    python -c "import telethon; import requests; import openpyxl; print('All dependencies OK')" 2>nul
    if errorlevel 1 (
        echo [ERROR] Dependency verification failed
        echo [INFO] Trying to reinstall...
        python -m pip install telethon requests openpyxl --upgrade --force-reinstall
        if errorlevel 1 (
            echo [ERROR] Reinstallation failed
            pause
            exit /b 1
        )
    ) else (
        echo [SUCCESS] All dependencies installed successfully
    )
) else (
    REM Verify all dependencies can be imported
    python -c "import telethon; import requests; import openpyxl; print('OK')" >nul 2>&1
    if errorlevel 1 (
        echo [WARN] Some dependencies cannot be imported, reinstalling...
        python -m pip install telethon requests openpyxl --upgrade --force-reinstall
        if errorlevel 1 (
            echo [ERROR] Reinstallation failed
            pause
            exit /b 1
        )
    )
    echo [OK] All dependencies are available
)

REM Step 4: Setup environment
echo.
echo [4/4] Setting up environment...
REM Set environment variables
set "NODE_ID={request.node_id}"
set "SERVER_URL={request.server_url}"
set "HEARTBEAT_INTERVAL={request.heartbeat_interval}"
set "SESSIONS_DIR=%~dp0sessions"

REM Create sessions directory if not exists
if not exist "!SESSIONS_DIR!" (
    mkdir "!SESSIONS_DIR!"
    echo [OK] Created sessions directory
) else (
    echo [OK] Sessions directory exists
)

echo.
echo ========================================
echo Starting Worker Node
echo ========================================
echo Node ID: !NODE_ID!
echo Server: !SERVER_URL!
echo Sessions Dir: !SESSIONS_DIR!
echo.

REM Run Python script
python worker_client.py

if errorlevel 1 (
    echo.
    echo [ERROR] Worker node failed
    pause
)
"""

        # 生成 Linux 启动脚本
        linux_script = f"""#!/bin/bash
set -e

cd "$(dirname "$0")"

echo ""
echo "========================================"
echo "Worker Node Auto Deploy"
echo "Node ID: {request.node_id}"
echo "========================================"
echo ""

# Step 1: Check Python
echo "[1/4] Checking Python installation..."
if ! command -v python3 &> /dev/null; then
    echo "[ERROR] Python3 not found!"
    echo ""
    echo "Please install Python 3.9 or higher:"
    echo "  Ubuntu/Debian: sudo apt-get install python3 python3-pip"
    echo "  CentOS/RHEL: sudo yum install python3 python3-pip"
    echo "  macOS: brew install python3"
    echo ""
    exit 1
fi

PYTHON_VERSION=$(python3 --version 2>&1 | awk '{{print $2}}')
echo "[OK] Python found: $PYTHON_VERSION"

# Check Python version (need 3.9+)
PYTHON_MAJOR=$(echo $PYTHON_VERSION | cut -d. -f1)
PYTHON_MINOR=$(echo $PYTHON_VERSION | cut -d. -f2)
if [ "$PYTHON_MAJOR" -lt 3 ] || ([ "$PYTHON_MAJOR" -eq 3 ] && [ "$PYTHON_MINOR" -lt 9 ]); then
    echo "[ERROR] Python 3.9+ required, found: $PYTHON_VERSION"
    exit 1
fi

# Step 2: Check and upgrade pip
echo ""
echo "[2/4] Checking pip..."
if ! command -v pip3 &> /dev/null && ! python3 -m pip --version &> /dev/null; then
    echo "[WARN] pip not found, installing..."
    python3 -m ensurepip --upgrade
    if [ $? -ne 0 ]; then
        echo "[ERROR] Failed to install pip"
        echo "Please install pip manually: sudo apt-get install python3-pip"
        exit 1
    fi
fi

echo "[OK] pip is available"
echo "[INFO] Upgrading pip to latest version..."
python3 -m pip install --upgrade pip -i https://pypi.tuna.tsinghua.edu.cn/simple &>/dev/null || python3 -m pip install --upgrade pip &>/dev/null

# Step 3: Check and install dependencies
echo ""
echo "[3/4] Checking Python dependencies..."
MISSING_DEPS=0

check_dependency() {{
    local pkg=$1
    if python3 -c "import $pkg" &>/dev/null; then
        local version=$(python3 -c "import $pkg; print($pkg.__version__)" 2>/dev/null || echo "unknown")
        echo "[OK] $pkg: $version"
        return 0
    else
        echo "[MISSING] $pkg"
        MISSING_DEPS=1
        return 1
    fi
}}

check_dependency telethon
check_dependency requests
check_dependency openpyxl

if [ $MISSING_DEPS -eq 1 ]; then
    echo ""
    echo "[INSTALL] Installing missing dependencies..."
    echo "[INFO] Using Tsinghua mirror source (faster in China)..."
    python3 -m pip install telethon requests openpyxl -i https://pypi.tuna.tsinghua.edu.cn/simple
    if [ $? -ne 0 ]; then
        echo "[WARN] Mirror source failed, trying default source..."
        python3 -m pip install telethon requests openpyxl --upgrade
        if [ $? -ne 0 ]; then
            echo "[ERROR] Failed to install dependencies"
            echo ""
            echo "Troubleshooting:"
            echo "  1. Check your internet connection"
            echo "  2. Try: python3 -m pip install telethon requests openpyxl --upgrade"
            echo "  3. Check firewall/proxy settings"
            exit 1
        fi
    fi
    
    echo ""
    echo "[VERIFY] Verifying installation..."
    python3 -c "import telethon; import requests; import openpyxl; print('All dependencies OK')" || {{
        echo "[ERROR] Dependency verification failed"
        exit 1
    }}
    echo "[SUCCESS] All dependencies installed successfully"
else
    # Verify all dependencies can be imported
    python3 -c "import telethon; import requests; import openpyxl; print('OK')" &>/dev/null
    if [ $? -ne 0 ]; then
        echo "[WARN] Some dependencies cannot be imported, reinstalling..."
        python3 -m pip install telethon requests openpyxl --upgrade --force-reinstall
    fi
    echo "[OK] All dependencies are available"
fi

# Step 4: Setup environment
echo ""
echo "[4/4] Setting up environment..."

# 设置环境变量
export NODE_ID="{request.node_id}"
export SERVER_URL="{request.server_url}"
export HEARTBEAT_INTERVAL="{request.heartbeat_interval}"
export SESSIONS_DIR="$(pwd)/sessions"

# 确保 sessions 目录存在
mkdir -p "$SESSIONS_DIR"

echo "[2/3] 启动 Worker 节点..."
echo "Node ID: $NODE_ID"
echo "Server: $SERVER_URL"
echo ""

# 运行 Python 脚本
python3 worker_client.py
"""

        # 生成 Worker 客户端 Python 代码
        worker_client_code = f'''#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Worker 节点客户端
Node ID: {request.node_id}
Server: {request.server_url}
"""

import os
import sys
import time
import json
import logging
import requests
import sqlite3
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# 配置
NODE_ID = os.getenv("NODE_ID", "{request.node_id}")
SERVER_URL = os.getenv("SERVER_URL", "{request.server_url}").rstrip('/')
HEARTBEAT_INTERVAL = int(os.getenv("HEARTBEAT_INTERVAL", "{request.heartbeat_interval}"))
SESSIONS_DIR = Path(os.getenv("SESSIONS_DIR", "./sessions")).resolve()
DEFAULT_API_ID = os.getenv("TELEGRAM_API_ID", "{request.telegram_api_id}")
DEFAULT_API_HASH = os.getenv("TELEGRAM_API_HASH", "{request.telegram_api_hash}")

# 确保 sessions 目录存在
SESSIONS_DIR.mkdir(exist_ok=True)

# Excel 配置文件 - 在多个位置查找
def find_excel_file() -> Optional[Path]:
    """查找 Excel 配置文件，支持多个位置和智能匹配"""
    node_id = NODE_ID
    excel_name = node_id + ".xlsx"
    
    def check_file_in_dir(directory: Path, filename: str) -> Optional[Path]:
        """在指定目录中查找文件，支持多种文件名变体"""
        # 1. 尝试精确匹配
        file_path = directory / filename
        if file_path.exists():
            return file_path
        
        # 2. 尝试双重扩展名（Windows 常见问题：.xlsx.xlsx）
        double_ext_file = directory / (filename + ".xlsx")
        if double_ext_file.exists():
            logger.warning(f"[EXCEL] 找到双重扩展名文件: {{double_ext_file}}，使用此文件")
            return double_ext_file
        
        # 3. 列出目录中所有 .xlsx 文件，查找匹配的
        for xlsx_file in directory.glob("*.xlsx*"):
            # 检查文件名是否以 node_id 开头
            if xlsx_file.stem.startswith(node_id) or xlsx_file.name.startswith(node_id):
                logger.info(f"[EXCEL] 找到匹配的 Excel 文件: {{xlsx_file}}")
                return xlsx_file
        
        return None
    
    def match_excel_by_phone_numbers(directory: Path) -> Optional[Path]:
        """通过电话号码匹配 Excel 文件"""
        try:
            import openpyxl
            
            # 获取所有 session 文件的电话号码
            session_phones = set()
            if SESSIONS_DIR.exists():
                for session_file in SESSIONS_DIR.glob("*.session"):
                    # 从文件名提取电话号码（去掉 .session 扩展名）
                    phone = session_file.stem
                    session_phones.add(phone)
            
            if not session_phones:
                logger.debug("[EXCEL] 没有找到 session 文件，无法进行电话号码匹配")
                return None
            
            logger.debug(f"[EXCEL] 找到 {{len(session_phones)}} 个 session 文件的电话号码")
            
            # 扫描目录中的所有 Excel 文件
            xlsx_files = list(directory.glob("*.xlsx"))
            if not xlsx_files:
                return None
            
            logger.info(f"[EXCEL] 在目录 {{directory}} 中找到 {{len(xlsx_files)}} 个 Excel 文件，尝试通过电话号码匹配...")
            
            for xlsx_file in xlsx_files:
                try:
                    wb = openpyxl.load_workbook(xlsx_file, read_only=True)
                    ws = wb.active
                    
                    # 读取表头
                    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
                    phone_idx = next((i for i, h in enumerate(headers) if 'phone' in h), None)
                    
                    if phone_idx is None:
                        logger.debug(f"[EXCEL] {{xlsx_file.name}} 没有 phone 列，跳过")
                        wb.close()
                        continue
                    
                    # 读取 Excel 中的电话号码
                    excel_phones = set()
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[phone_idx]:
                            phone = str(row[phone_idx]).strip()
                            if phone:
                                excel_phones.add(phone)
                    
                    wb.close()
                    
                    # 检查是否有匹配的电话号码
                    matched_phones = session_phones & excel_phones
                    if matched_phones:
                        match_count = len(matched_phones)
                        total_excel = len(excel_phones)
                        total_session = len(session_phones)
                        logger.info(f"[EXCEL] 找到匹配的 Excel 文件: {{xlsx_file.name}}")
                        logger.info(f"[EXCEL] 匹配详情: Excel 中有 {{total_excel}} 个号码，Session 中有 {{total_session}} 个号码，匹配了 {{match_count}} 个")
                        return xlsx_file
                    else:
                        logger.debug(f"[EXCEL] {{xlsx_file.name}} 中的电话号码与 session 文件不匹配")
                        
                except Exception as e:
                    logger.debug(f"[EXCEL] 读取 {{xlsx_file.name}} 失败: {{e}}")
                    continue
            
            return None
            
        except ImportError:
            logger.debug("[EXCEL] openpyxl 未安装，无法进行电话号码匹配")
            return None
        except Exception as e:
            logger.debug(f"[EXCEL] 电话号码匹配过程出错: {{e}}")
            return None
    
    # 1. 当前工作目录
    current_dir = Path.cwd()
    excel_file = check_file_in_dir(current_dir, excel_name)
    if excel_file:
        logger.info(f"[EXCEL] 在当前目录找到: {{excel_file}}")
        return excel_file
    
    # 2. 脚本所在目录（worker_client.py 所在目录）
    # 使用 __file__ 获取脚本路径（更可靠）
    try:
        # 尝试使用 __file__（如果在 worker_client.py 中运行）
        script_dir = Path(__file__).resolve().parent
    except NameError:
        # 如果 __file__ 不存在，使用 sys.argv[0]
        import sys
        if len(sys.argv) > 0 and sys.argv[0]:
            script_path = Path(sys.argv[0])
            # 如果路径不是绝对路径，转换为绝对路径
            if not script_path.is_absolute():
                script_path = Path.cwd() / script_path
            script_dir = script_path.resolve().parent
        else:
            script_dir = Path.cwd()
    
    # 2. 脚本所在目录（使用改进的查找函数）
    excel_file = check_file_in_dir(script_dir, excel_name)
    if excel_file:
        logger.info(f"[EXCEL] 在脚本目录找到: {{excel_file}}")
        return excel_file
    
    # 3. sessions 目录的父目录（如果 sessions 是相对路径）
    if not SESSIONS_DIR.is_absolute():
        parent_dir = SESSIONS_DIR.parent
        excel_file = check_file_in_dir(parent_dir, excel_name)
        if excel_file:
            logger.info(f"[EXCEL] 在 sessions 父目录找到: {{excel_file}}")
            return excel_file
    else:
        # sessions 是绝对路径，检查父目录
        parent_dir = SESSIONS_DIR.parent
        excel_file = check_file_in_dir(parent_dir, excel_name)
        if excel_file:
            logger.info(f"[EXCEL] 在 sessions 父目录找到: {{excel_file}}")
            return excel_file
    
    # 4. sessions 目录本身
    excel_file = check_file_in_dir(SESSIONS_DIR, excel_name)
    if excel_file:
        logger.info(f"[EXCEL] 在 sessions 目录找到: {{excel_file}}")
        return excel_file
    
    # 5. 智能匹配：通过电话号码匹配脚本目录中的 Excel 文件
    logger.info(f"[EXCEL] 未找到名称匹配的配置文件，尝试通过电话号码智能匹配...")
    excel_file = match_excel_by_phone_numbers(script_dir)
    if excel_file:
        logger.info(f"[EXCEL] 通过电话号码匹配找到 Excel 文件: {{excel_file}}")
        return excel_file
    
    # 6. 智能匹配：通过电话号码匹配当前工作目录中的 Excel 文件
    excel_file = match_excel_by_phone_numbers(current_dir)
    if excel_file:
        logger.info(f"[EXCEL] 通过电话号码匹配找到 Excel 文件: {{excel_file}}")
        return excel_file
    
    # 未找到 - 输出详细的调试信息
    logger.warning(f"[EXCEL] 未找到配置文件，查找了以下位置：")
    logger.warning(f"  1. 当前工作目录: {{current_dir / excel_name}}")
    logger.warning(f"  2. 脚本所在目录: {{script_dir / excel_name}}")
    logger.warning(f"  3. sessions 父目录: {{SESSIONS_DIR.parent / excel_name}}")
    logger.warning(f"  4. sessions 目录: {{SESSIONS_DIR / excel_name}}")
    logger.warning(f"[DEBUG] 当前工作目录: {{Path.cwd()}}")
    logger.warning(f"[DEBUG] 脚本目录: {{script_dir}}")
    logger.warning(f"[DEBUG] sessions 目录: {{SESSIONS_DIR}}")
    logger.warning(f"[DEBUG] Excel 文件名: {{excel_name}}")
    logger.warning(f"[DEBUG] NODE_ID: {{NODE_ID}}")
    
    # 额外尝试：列出脚本目录中的所有 .xlsx 文件
    try:
        xlsx_files = list(script_dir.glob("*.xlsx"))
        if xlsx_files:
            logger.info(f"[DEBUG] 脚本目录中的 Excel 文件: {{[f.name for f in xlsx_files]}}")
    except Exception:
        pass
    
    return None

# Excel 配置文件
EXCEL_FILE = find_excel_file()


def load_accounts_from_excel() -> List[Dict[str, Any]]:
    """从 Excel 文件加载账号配置"""
    accounts = []
    
    # 重新查找 Excel 文件（每次调用都查找，以防文件移动）
    excel_file = find_excel_file()
    if not excel_file or not excel_file.exists():
        excel_name = NODE_ID + ".xlsx"  # 使用变量拼接，避免字面量问题
        if EXCEL_FILE:
            logger.warning(f"Excel 配置文件不存在: {{EXCEL_FILE}}")
        else:
            logger.warning(f"Excel 配置文件不存在: {{excel_name}}")
        return accounts
    
    # 使用找到的文件
    excel_file_to_use = excel_file
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_file_to_use)
        ws = wb.active
        logger.info(f"[EXCEL] 正在读取文件: {{excel_file_to_use}}")
        
        # 读取表头
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
        
        # 查找列索引
        api_id_idx = next((i for i, h in enumerate(headers) if 'api_id' in h), None)
        api_hash_idx = next((i for i, h in enumerate(headers) if 'api_hash' in h), None)
        phone_idx = next((i for i, h in enumerate(headers) if 'phone' in h), None)
        enabled_idx = next((i for i, h in enumerate(headers) if 'enabled' in h), None)
        
        if api_id_idx is None or api_hash_idx is None or phone_idx is None:
            logger.error(f"Excel 文件缺少必需的列：需要 api_id, api_hash, phone")
            return accounts
        
        # 读取数据行
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):
                continue
            
            phone = str(row[phone_idx]).strip() if row[phone_idx] else None
            api_id = str(row[api_id_idx]).strip() if row[api_id_idx] else DEFAULT_API_ID
            api_hash = str(row[api_hash_idx]).strip() if row[api_hash_idx] else DEFAULT_API_HASH
            enabled = True
            if enabled_idx is not None and row[enabled_idx] is not None:
                enabled = bool(int(row[enabled_idx])) if str(row[enabled_idx]).isdigit() else bool(row[enabled_idx])
            
            if phone and api_id and api_hash and enabled:
                accounts.append({{
                    "phone": phone,
                    "api_id": int(api_id) if api_id.isdigit() else None,
                    "api_hash": api_hash,
                    "enabled": enabled
                }})
        
        logger.info(f"从 Excel 加载了 {{len(accounts)}} 个账号")
        return accounts
        
    except ImportError:
        logger.error("openpyxl 未安装，无法读取 Excel 文件")
        logger.info("安装命令: pip install openpyxl")
        return accounts
    except Exception as e:
        logger.error(f"读取 Excel 文件失败: {{e}}", exc_info=True)
        return accounts


def scan_session_files() -> List[str]:
    """扫描 sessions 目录中的 .session 文件"""
    session_files = []
    
    if not SESSIONS_DIR.exists():
        return session_files
    
    for file_path in SESSIONS_DIR.glob("*.session"):
        session_files.append(file_path.name)
    
    logger.info(f"扫描到 {{len(session_files)}} 个 session 文件")
    return session_files


def get_account_info_via_telethon(phone: str, api_id: Optional[int], api_hash: Optional[str]) -> Dict[str, Any]:
    """通过 Telethon 获取账号信息"""
    try:
        from telethon import TelegramClient
        from telethon.errors import SessionPasswordNeeded, PhoneCodeInvalid
        
        session_file = SESSIONS_DIR / f"{{phone}}.session"
        if not session_file.exists():
            return {{"phone": phone, "user_id": None, "username": None, "name": None}}
        
        # 使用 Excel 中的 API ID/Hash，如果没有则使用默认值
        effective_api_id = api_id or (int(DEFAULT_API_ID) if DEFAULT_API_ID.isdigit() else None)
        effective_api_hash = api_hash or DEFAULT_API_HASH
        
        if not effective_api_id or not effective_api_hash:
            logger.warning(f"账号 {{phone}} 缺少 API_ID 或 API_HASH，跳过 Telethon 查询")
            return {{"phone": phone, "user_id": None, "username": None, "name": None}}
        
        client = TelegramClient(str(session_file), effective_api_id, effective_api_hash)
        
        try:
            client.connect()
            
            if not client.is_connected():
                return {{"phone": phone, "user_id": None, "username": None, "name": None}}
            
            me = client.get_me()
            if me:
                return {{
                    "phone": phone,
                    "user_id": me.id,
                    "username": me.username or "",
                    "name": f"{{me.first_name or ''}} {{me.last_name or ''}}".strip() or "",
                    "status": "online"
                }}
        finally:
            client.disconnect()
            
    except ImportError:
        logger.warning("Telethon 未安装，无法获取账号详细信息")
    except Exception as e:
        logger.debug(f"获取账号 {{phone}} 信息失败: {{e}}")
    
    return {{"phone": phone, "user_id": None, "username": None, "name": None, "status": "offline"}}


def prepare_accounts_for_heartbeat() -> List[Dict[str, Any]]:
    """准备发送给服务器的账号列表"""
    accounts_data = []
    
    # 从 Excel 加载账号配置
    excel_accounts = load_accounts_from_excel()
    
    # 扫描 session 文件
    session_files = scan_session_files()
    
    # 匹配 Excel 配置和 Session 文件
    for excel_acc in excel_accounts:
        phone = excel_acc["phone"]
        session_file = SESSIONS_DIR / f"{{phone}}.session"
        
        if session_file.exists():
            # 尝试通过 Telethon 获取详细信息
            account_info = get_account_info_via_telethon(
                phone,
                excel_acc.get("api_id"),
                excel_acc.get("api_hash")
            )
            
            # 合并 Excel 配置和 Telethon 信息
            account_data = {{
                "account_id": phone,
                "phone": phone,
                "user_id": account_info.get("user_id"),
                "username": account_info.get("username"),
                "name": account_info.get("name"),
                "status": account_info.get("status", "offline"),
                "api_id": excel_acc.get("api_id"),
                "api_hash": excel_acc.get("api_hash")[:10] + "..." if excel_acc.get("api_hash") else None,  # 只显示部分
            }}
            accounts_data.append(account_data)
        else:
            # Session 文件不存在，但仍然添加到列表中（状态为 offline）
            accounts_data.append({{
                "account_id": phone,
                "phone": phone,
                "status": "offline",
                "api_id": excel_acc.get("api_id"),
            }})
    
    return accounts_data


def send_heartbeat() -> bool:
    """发送心跳到服务器"""
    try:
        accounts = prepare_accounts_for_heartbeat()
        
        heartbeat_data = {{
            "node_id": NODE_ID,
            "status": "online",
            "account_count": len(accounts),
            "accounts": accounts,
            "metadata": {{
                "heartbeat_interval": HEARTBEAT_INTERVAL,
                "sessions_dir": str(SESSIONS_DIR),
                "timestamp": datetime.now().isoformat()
            }}
        }}
        
        response = requests.post(
            f"{{SERVER_URL}}/api/v1/workers/heartbeat",
            json=heartbeat_data,
            timeout=10,
            headers={{'Content-Type': 'application/json'}}
        )
        
        if response.ok:
            data = response.json()
            commands = data.get("pending_commands", [])
            if commands:
                logger.info(f"收到 {{len(commands)}} 个待执行命令")
            return True
        else:
            logger.warning(f"心跳失败: HTTP {{response.status_code}}")
            return False
            
    except requests.exceptions.RequestException as e:
        logger.error(f"发送心跳异常: {{e}}")
        return False
    except Exception as e:
        logger.error(f"准备心跳数据失败: {{e}}", exc_info=True)
        return False


def main():
    """主函数"""
    logger.info("=" * 60)
    logger.info(f"Worker Node: {{NODE_ID}}")
    logger.info(f"Server: {{SERVER_URL}}")
    logger.info(f"Sessions Dir: {{SESSIONS_DIR}}")
    logger.info(f"Heartbeat Interval: {{HEARTBEAT_INTERVAL}}s")
    logger.info("=" * 60)
    logger.info("")
    
    logger.info("[INIT] 初始化 Worker 节点...")
    
    # 检查 Excel 文件
    excel_file = find_excel_file()
    excel_name = NODE_ID + ".xlsx"  # 定义 excel_name 用于日志输出
    if excel_file and excel_file.exists():
        logger.info(f"[EXCEL] 找到配置文件: {{excel_file}}")
        logger.info(f"[EXCEL] 文件路径: {{excel_file.absolute()}}")
    else:
        logger.warning(f"[EXCEL] 配置文件不存在: {{excel_name}}")
        logger.info("请创建 Excel 配置文件，包含以下列：api_id, api_hash, phone, enabled")
        logger.info(f"[EXCEL] 建议放置在以下位置之一：")
        logger.info(f"  - {{Path.cwd() / excel_name}}")
        logger.info(f"  - {{SESSIONS_DIR.parent / excel_name}}")
    
    # 扫描 Session 文件
    session_files = scan_session_files()
    logger.info(f"[SESSION] 找到 {{len(session_files)}} 个 session 文件")
    
    # 心跳循环
    logger.info("[WORKER] 开始心跳循环...")
    try:
        while True:
            send_heartbeat()
            time.sleep(HEARTBEAT_INTERVAL)
    except KeyboardInterrupt:
        logger.info("[WORKER] 收到中断信号，正在停止...")
        send_heartbeat()  # 发送最后一次心跳
    except Exception as e:
        logger.error(f"[WORKER] 心跳循环异常: {{e}}", exc_info=True)


if __name__ == "__main__":
    main()
'''

        # 生成 fix_session 脚本
        fix_session_code = '''#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
修复 Telethon Session 文件版本问题
使用方法: python fix_session.py [sessions_directory]
"""

import sys
import sqlite3
from pathlib import Path

def fix_session_version(session_file: Path):
    """修复 session 文件的 version 列问题"""
    try:
        conn = sqlite3.connect(str(session_file))
        cursor = conn.cursor()
        
        # 检查表结构
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='sessions'")
        if not cursor.fetchone():
            print(f"  {session_file.name}: sessions 表不存在，跳过")
            conn.close()
            return False
        
        # 检查是否有 version 列
        cursor.execute("PRAGMA table_info(sessions)")
        columns = [col[1] for col in cursor.fetchall()]
        
        if 'version' not in columns:
            print(f"  {session_file.name}: 添加 version 列...")
            cursor.execute("ALTER TABLE sessions ADD COLUMN version INTEGER DEFAULT 1")
            conn.commit()
            print(f"  {session_file.name}: ✅ 已修复")
        else:
            print(f"  {session_file.name}: version 列已存在，跳过")
        
        conn.close()
        return True
        
    except Exception as e:
        print(f"  {session_file.name}: ❌ 修复失败: {e}")
        return False

def main():
    sessions_dir = Path(sys.argv[1] if len(sys.argv) > 1 else "./sessions")
    
    if not sessions_dir.exists():
        print(f"错误: 目录不存在: {sessions_dir}")
        sys.exit(1)
    
    print(f"扫描目录: {sessions_dir}")
    session_files = list(sessions_dir.glob("*.session"))
    
    if not session_files:
        print("未找到 .session 文件")
        sys.exit(0)
    
    print(f"找到 {len(session_files)} 个 session 文件")
    print("")
    
    fixed = 0
    for session_file in session_files:
        if fix_session_version(session_file):
            fixed += 1
    
    print("")
    print(f"修复完成: {fixed}/{len(session_files)} 个文件")

if __name__ == "__main__":
    main()
'''

        # 生成 create_excel_template 脚本
        create_excel_template_code = f'''#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
创建 Excel 配置文件模板
使用方法: python create_excel_template.py
"""

import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    print("错误: openpyxl 未安装")
    print("安装命令: pip install openpyxl")
    sys.exit(1)

def main():
    node_id = "{request.node_id}"
    excel_file = Path(f"{{node_id}}.xlsx")
    
    if excel_file.exists():
        response = input(f"文件 {{excel_file}} 已存在，是否覆盖？(y/N): ")
        if response.lower() != 'y':
            print("已取消")
            sys.exit(0)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Accounts"
    
    # 写入表头
    headers = ["api_id", "api_hash", "phone", "enabled"]
    ws.append(headers)
    
    # 写入示例数据（将被用户替换）
    ws.append(["YOUR_API_ID", "YOUR_API_HASH", "639123456789", "1"])
    
    # 保存文件
    wb.save(excel_file)
    print(f"✅ Excel 模板已创建: {{excel_file}}")
    print("")
    print("请编辑此文件，添加您的账号信息：")
    print("  - api_id: Telegram API ID（从 my.telegram.org 获取）")
    print("  - api_hash: Telegram API Hash（从 my.telegram.org 获取）")
    print("  - phone: 电话号码（必须与 session 文件名匹配）")
    print("  - enabled: 1=启用，0=禁用")
    print("")
    print("注意：Session 文件名必须是 {{phone}}.session 格式")

if __name__ == "__main__":
    main()
'''

        return {
            "success": True,
            "scripts": {
                "windows": windows_script,
                "linux": linux_script,
                "worker_client": worker_client_code,
                "fix_session": fix_session_code,
                "create_excel_template": create_excel_template_code
            },
            "message": "部署包生成成功"
        }
        
    except Exception as e:
        logger.error(f"生成部署包失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"生成部署包失败: {str(e)}"
        )


# ============ Session 管理 API ============

@router.get("/{worker_id}/sessions", response_model=SessionListResponse, status_code=status.HTTP_200_OK)
async def get_worker_sessions(
    worker_id: str,
    timeout: int = 30,
    current_user: Optional[User] = Depends(get_current_active_user),
    db: Session = Depends(get_db_session)
):
    """
    获取 Worker 节点的 Session 文件列表
    
    工作流程：
    1. 检查 Worker 节点是否在线
    2. 生成唯一的命令 ID
    3. 通过命令队列发送 "list_sessions" 命令
    4. 等待 Worker 节点响应（通过心跳返回结果）
    5. 返回 Session 文件列表
    """
    try:
        # 检查 Worker 节点是否在线
        worker_status = _get_worker_status(worker_id)
        if not worker_status:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Worker 节点 {worker_id} 不存在"
            )
        
        if worker_status.get("status") != "online":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Worker 节点 {worker_id} 不在线"
            )
        
        # 生成唯一的命令 ID
        import uuid
        command_id = str(uuid.uuid4())
        
        # 发送命令到 Worker 节点
        command = {
            "action": "list_sessions",
            "params": {},
            "command_id": command_id,
            "timestamp": datetime.now().isoformat(),
            "from": "master"
        }
        
        _add_command(worker_id, command)
        logger.info(f"向节点 {worker_id} 发送 list_sessions 命令 (ID: {command_id})")
        
        # 等待响应（轮询方式）
        response = _wait_for_response(worker_id, command_id, timeout=timeout)
        
        if not response:
            raise HTTPException(
                status_code=status.HTTP_504_GATEWAY_TIMEOUT,
                detail=f"等待 Worker 节点 {worker_id} 响应超时（{timeout}秒）"
            )
        
        if not response.get("success"):
            error_msg = response.get("error", "未知错误")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Worker 节点执行失败: {error_msg}"
            )
        
        # 解析响应数据
        sessions_data = response.get("data", {}).get("sessions", [])
        sessions = []
        for session_info in sessions_data:
            sessions.append(SessionFileInfo(
                filename=session_info.get("filename", ""),
                size=session_info.get("size", 0),
                modified_time=session_info.get("modified_time"),
                path=session_info.get("path")
            ))
        
        return SessionListResponse(
            success=True,
            node_id=worker_id,
            sessions=sessions,
            total=len(sessions),
            message=f"成功获取 {len(sessions)} 个 Session 文件"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取 Worker Session 列表失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取 Session 列表失败: {str(e)}"
        )


@router.post("/{worker_id}/sessions/upload", response_model=SessionUploadResponse, status_code=status.HTTP_200_OK)
async def upload_session_to_worker(
    worker_id: str,
    file: UploadFile = File(...),
    timeout: int = 60,
    current_user: Optional[User] = Depends(get_current_active_user),
    db: Session = Depends(get_db_session)
):
    """
    上传 Session 文件到指定的 Worker 节点
    
    工作流程：
    1. 检查 Worker 节点是否在线
    2. 验证文件格式（必须是 .session 文件）
    3. 读取文件内容并编码为 base64
    4. 生成唯一的命令 ID
    5. 通过命令队列发送 "upload_session" 命令，包含文件内容
    6. 等待 Worker 节点响应
    7. 返回上传结果
    """
    try:
        # 检查 Worker 节点是否在线
        worker_status = _get_worker_status(worker_id)
        if not worker_status:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Worker 节点 {worker_id} 不存在"
            )
        
        if worker_status.get("status") != "online":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Worker 节点 {worker_id} 不在线"
            )
        
        # 验证文件格式
        if not file.filename or not file.filename.endswith('.session'):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="只支持 .session 文件"
            )
        
        # 读取文件内容
        file_content = await file.read()
        file_size = len(file_content)
        
        # 检查文件大小（限制为 10MB）
        MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
        if file_size > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"文件大小超过限制（最大 10MB），当前大小: {file_size / 1024 / 1024:.2f}MB"
            )
        
        # 编码为 base64
        file_content_b64 = base64.b64encode(file_content).decode('utf-8')
        
        # 生成唯一的命令 ID
        import uuid
        command_id = str(uuid.uuid4())
        
        # 发送命令到 Worker 节点
        command = {
            "action": "upload_session",
            "params": {
                "filename": file.filename,
                "file_content": file_content_b64,
                "file_size": file_size
            },
            "command_id": command_id,
            "timestamp": datetime.now().isoformat(),
            "from": "master"
        }
        
        _add_command(worker_id, command)
        logger.info(f"向节点 {worker_id} 发送 upload_session 命令 (ID: {command_id}, 文件: {file.filename}, 大小: {file_size} bytes)")
        
        # 等待响应（轮询方式）
        response = _wait_for_response(worker_id, command_id, timeout=timeout)
        
        if not response:
            raise HTTPException(
                status_code=status.HTTP_504_GATEWAY_TIMEOUT,
                detail=f"等待 Worker 节点 {worker_id} 响应超时（{timeout}秒）"
            )
        
        if not response.get("success"):
            error_msg = response.get("error", "未知错误")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Worker 节点上传失败: {error_msg}"
            )
        
        # 解析响应数据
        result_data = response.get("data", {})
        saved_filename = result_data.get("filename", file.filename)
        
        return SessionUploadResponse(
            success=True,
            node_id=worker_id,
            filename=saved_filename,
            message=f"Session 文件已成功上传到节点 {worker_id}"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"上传 Session 文件到 Worker 失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"上传 Session 文件失败: {str(e)}"
        )


@router.get("/{worker_id}/sessions/{filename}/download", response_model=SessionDownloadResponse, status_code=status.HTTP_200_OK)
async def download_session_from_worker(
    worker_id: str,
    filename: str,
    timeout: int = 30,
    current_user: Optional[User] = Depends(get_current_active_user),
    db: Session = Depends(get_db_session)
):
    """
    从 Worker 节点下载 Session 文件
    
    工作流程：
    1. 检查 Worker 节点是否在线
    2. 生成唯一的命令 ID
    3. 通过命令队列发送 "download_session" 命令
    4. 等待 Worker 节点响应
    5. 返回 base64 编码的文件内容
    """
    try:
        # 检查 Worker 节点是否在线
        worker_status = _get_worker_status(worker_id)
        if not worker_status:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Worker 节点 {worker_id} 不存在"
            )
        
        if worker_status.get("status") != "online":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Worker 节点 {worker_id} 不在线"
            )
        
        # 验证文件名
        if not filename or not filename.endswith('.session'):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="只支持 .session 文件"
            )
        
        # 生成唯一的命令 ID
        import uuid
        command_id = str(uuid.uuid4())
        
        # 发送命令到 Worker 节点
        command = {
            "action": "download_session",
            "params": {
                "filename": filename
            },
            "command_id": command_id,
            "timestamp": datetime.now().isoformat(),
            "from": "master"
        }
        
        _add_command(worker_id, command)
        logger.info(f"向节点 {worker_id} 发送 download_session 命令 (ID: {command_id}, 文件: {filename})")
        
        # 等待响应（轮询方式）
        response = _wait_for_response(worker_id, command_id, timeout=timeout)
        
        if not response:
            raise HTTPException(
                status_code=status.HTTP_504_GATEWAY_TIMEOUT,
                detail=f"等待 Worker 节点 {worker_id} 响应超时（{timeout}秒）"
            )
        
        if not response.get("success"):
            error_msg = response.get("error", "未知错误")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Worker 节点下载失败: {error_msg}"
            )
        
        # 解析响应数据
        result_data = response.get("data", {})
        file_content_b64 = result_data.get("file_content", "")
        file_size = result_data.get("file_size", 0)
        
        if not file_content_b64:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Worker 节点返回的文件内容为空"
            )
        
        return SessionDownloadResponse(
            success=True,
            node_id=worker_id,
            filename=filename,
            file_content=file_content_b64,
            file_size=file_size,
            message=f"Session 文件已成功从节点 {worker_id} 下载"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"下载 Session 文件失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"下载 Session 文件失败: {str(e)}"
        )


@router.delete("/{worker_id}/sessions/{filename}", response_model=SessionDeleteResponse, status_code=status.HTTP_200_OK)
async def delete_session_from_worker(
    worker_id: str,
    filename: str,
    timeout: int = 30,
    current_user: Optional[User] = Depends(get_current_active_user),
    db: Session = Depends(get_db_session)
):
    """
    从 Worker 节点删除 Session 文件
    
    工作流程：
    1. 检查 Worker 节点是否在线
    2. 生成唯一的命令 ID
    3. 通过命令队列发送 "delete_session" 命令
    4. 等待 Worker 节点响应
    5. 返回删除结果
    """
    try:
        # 检查 Worker 节点是否在线
        worker_status = _get_worker_status(worker_id)
        if not worker_status:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Worker 节点 {worker_id} 不存在"
            )
        
        if worker_status.get("status") != "online":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Worker 节点 {worker_id} 不在线"
            )
        
        # 验证文件名
        if not filename or not filename.endswith('.session'):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="只支持 .session 文件"
            )
        
        # 生成唯一的命令 ID
        import uuid
        command_id = str(uuid.uuid4())
        
        # 发送命令到 Worker 节点
        command = {
            "action": "delete_session",
            "params": {
                "filename": filename
            },
            "command_id": command_id,
            "timestamp": datetime.now().isoformat(),
            "from": "master"
        }
        
        _add_command(worker_id, command)
        logger.info(f"向节点 {worker_id} 发送 delete_session 命令 (ID: {command_id}, 文件: {filename})")
        
        # 等待响应（轮询方式）
        response = _wait_for_response(worker_id, command_id, timeout=timeout)
        
        if not response:
            raise HTTPException(
                status_code=status.HTTP_504_GATEWAY_TIMEOUT,
                detail=f"等待 Worker 节点 {worker_id} 响应超时（{timeout}秒）"
            )
        
        if not response.get("success"):
            error_msg = response.get("error", "未知错误")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Worker 节点删除失败: {error_msg}"
            )
        
        return SessionDeleteResponse(
            success=True,
            node_id=worker_id,
            filename=filename,
            message=f"Session 文件已成功从节点 {worker_id} 删除"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除 Session 文件失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"删除 Session 文件失败: {str(e)}"
        )


@router.post("/{worker_id}/sessions/batch", response_model=BatchOperationResponse, status_code=status.HTTP_200_OK)
async def batch_operation_sessions(
    worker_id: str,
    request: BatchOperationRequest,
    timeout: int = 120,
    current_user: Optional[User] = Depends(get_current_active_user),
    db: Session = Depends(get_db_session)
):
    """
    批量操作 Session 文件（下载、删除）
    
    注意：批量操作会为每个文件发送单独的命令，并等待所有响应
    """
    try:
        # 检查 Worker 节点是否在线
        worker_status = _get_worker_status(worker_id)
        if not worker_status:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Worker 节点 {worker_id} 不存在"
            )
        
        if worker_status.get("status") != "online":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Worker 节点 {worker_id} 不在线"
            )
        
        # 验证操作类型
        if request.operation not in ["download", "delete"]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"不支持的批量操作类型: {request.operation}。支持: download, delete"
            )
        
        # 批量发送命令
        import uuid
        command_ids = []
        for filename in request.filenames:
            command_id = str(uuid.uuid4())
            command_ids.append((filename, command_id))
            
            command = {
                "action": f"{request.operation}_session",
                "params": {
                    "filename": filename
                },
                "command_id": command_id,
                "timestamp": datetime.now().isoformat(),
                "from": "master"
            }
            
            _add_command(worker_id, command)
        
        logger.info(f"向节点 {worker_id} 发送批量 {request.operation} 命令: {len(request.filenames)} 个文件")
        
        # 等待所有响应
        results = []
        successful = 0
        failed = 0
        
        for filename, command_id in command_ids:
            response = _wait_for_response(worker_id, command_id, timeout=timeout)
            
            if response and response.get("success"):
                successful += 1
                results.append({
                    "filename": filename,
                    "success": True,
                    "data": response.get("data", {})
                })
            else:
                failed += 1
                error_msg = response.get("error", "超时或未知错误") if response else "超时"
                results.append({
                    "filename": filename,
                    "success": False,
                    "error": error_msg
                })
        
        return BatchOperationResponse(
            success=successful > 0,
            node_id=worker_id,
            operation=request.operation,
            total=len(request.filenames),
            successful=successful,
            failed=failed,
            results=results,
            message=f"批量操作完成: {successful} 成功, {failed} 失败"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"批量操作 Session 文件失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"批量操作失败: {str(e)}"
        )
