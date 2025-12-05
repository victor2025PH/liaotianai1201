"""
群組管理模組
支持：
- 自動建群
- 邀請其他 AI 帳號加入
- 未來：邀請真實用戶
"""

import asyncio
import random
import logging
from datetime import datetime
from typing import Optional, List, Dict, Any

from telethon import TelegramClient
from telethon.tl.functions.messages import (
    CreateChatRequest,
    AddChatUserRequest,
    ExportChatInviteRequest
)
from telethon.tl.functions.channels import (
    CreateChannelRequest,
    InviteToChannelRequest,
    EditPhotoRequest
)
from telethon.tl.types import (
    Chat, Channel, User,
    InputPeerUser, InputPeerChannel, InputPeerChat
)

logger = logging.getLogger(__name__)


class GroupManager:
    """群組管理器"""
    
    def __init__(self, client: TelegramClient, user_id: int):
        self.client = client
        self.user_id = user_id
        self.created_groups: List[dict] = []
    
    async def create_group(
        self,
        title: str,
        about: str = "",
        member_user_ids: List[int] = None,
        is_supergroup: bool = False
    ) -> Optional[dict]:
        """
        創建群組
        
        Args:
            title: 群組名稱
            about: 群組描述（僅超級群組）
            member_user_ids: 初始成員的 user_id 列表
            is_supergroup: 是否創建超級群組
        
        Returns:
            群組信息字典
        """
        try:
            if is_supergroup:
                # 創建超級群組（頻道轉群組）
                result = await self.client(CreateChannelRequest(
                    title=title,
                    about=about,
                    megagroup=True  # 這使它成為超級群組而不是頻道
                ))
                
                chat = result.chats[0]
                group_id = chat.id
                group_type = "supergroup"
                
                logger.info(f"[用戶 {self.user_id}] 創建超級群組成功: {title} (ID: {group_id})")
                
            else:
                # 創建普通群組
                users = []
                if member_user_ids:
                    for uid in member_user_ids[:5]:  # 創建時最多加5人
                        try:
                            user = await self.client.get_entity(uid)
                            if isinstance(user, User):
                                users.append(user)
                        except Exception as e:
                            logger.warning(f"無法獲取用戶 {uid}: {e}")
                
                # 至少需要一個用戶來創建群組
                if not users:
                    # 添加自己（創建空群後再刪除）
                    me = await self.client.get_me()
                    users = [me]
                
                result = await self.client(CreateChatRequest(
                    title=title,
                    users=users
                ))
                
                chat = result.chats[0]
                group_id = chat.id
                group_type = "group"
                
                logger.info(f"[用戶 {self.user_id}] 創建普通群組成功: {title} (ID: {group_id})")
            
            # 獲取邀請鏈接
            invite_link = await self.get_invite_link(chat)
            
            group_info = {
                "id": group_id,
                "title": title,
                "type": group_type,
                "invite_link": invite_link,
                "creator_id": self.user_id,
                "created_at": datetime.now().isoformat()
            }
            
            self.created_groups.append(group_info)
            return group_info
            
        except Exception as e:
            logger.error(f"[用戶 {self.user_id}] 創建群組失敗: {e}")
            return None
    
    async def get_invite_link(self, chat) -> Optional[str]:
        """獲取群組邀請鏈接"""
        try:
            if isinstance(chat, Channel):
                result = await self.client(ExportChatInviteRequest(
                    peer=InputPeerChannel(chat.id, chat.access_hash)
                ))
            else:
                result = await self.client(ExportChatInviteRequest(
                    peer=InputPeerChat(chat.id)
                ))
            
            return result.link
        except Exception as e:
            logger.error(f"獲取邀請鏈接失敗: {e}")
            return None
    
    async def invite_users(
        self,
        chat_id: int,
        user_ids: List[int],
        is_channel: bool = False
    ) -> Dict[int, bool]:
        """
        邀請用戶加入群組
        
        Returns:
            {user_id: success} 字典
        """
        results = {}
        
        for user_id in user_ids:
            try:
                user = await self.client.get_entity(user_id)
                
                if is_channel:
                    chat = await self.client.get_entity(chat_id)
                    await self.client(InviteToChannelRequest(
                        channel=chat,
                        users=[user]
                    ))
                else:
                    await self.client(AddChatUserRequest(
                        chat_id=chat_id,
                        user_id=user,
                        fwd_limit=50
                    ))
                
                results[user_id] = True
                logger.info(f"[用戶 {self.user_id}] 成功邀請 {user_id} 加入群組 {chat_id}")
                
                # 間隔避免限流
                await asyncio.sleep(random.uniform(2, 5))
                
            except Exception as e:
                results[user_id] = False
                logger.error(f"[用戶 {self.user_id}] 邀請 {user_id} 失敗: {e}")
        
        return results


class TestGroupOrchestrator:
    """測試群組協調器 - 協調多個帳號建群和加入"""
    
    def __init__(self):
        self.accounts: Dict[int, dict] = {}  # user_id -> {client, manager, config}
        self.test_group: Optional[dict] = None
    
    def add_account(
        self,
        user_id: int,
        client: TelegramClient,
        config: dict
    ):
        """添加帳號"""
        manager = GroupManager(client, user_id)
        self.accounts[user_id] = {
            "client": client,
            "manager": manager,
            "config": config,
            "username": config.get("username", ""),
            "phone": config.get("phone", "")
        }
        logger.info(f"添加帳號: {user_id} ({config.get('phone', 'N/A')})")
    
    async def create_test_group(
        self,
        creator_user_id: int = None,
        group_name: str = None
    ) -> Optional[dict]:
        """
        創建測試群組
        
        Args:
            creator_user_id: 指定創建者的 user_id，None 則選第一個帳號
            group_name: 群組名稱，None 則自動生成
        """
        if not self.accounts:
            logger.error("沒有可用帳號")
            return None
        
        # 選擇創建者
        if creator_user_id and creator_user_id in self.accounts:
            creator_id = creator_user_id
        else:
            creator_id = list(self.accounts.keys())[0]
        
        creator_account = self.accounts[creator_id]
        manager: GroupManager = creator_account["manager"]
        
        # 生成群組名稱
        if not group_name:
            timestamp = datetime.now().strftime("%m%d%H%M")
            group_name = f"🧧 紅包測試群 {timestamp}"
        
        # 獲取其他帳號的 user_id
        other_user_ids = [uid for uid in self.accounts.keys() if uid != creator_id]
        
        logger.info(f"創建者: {creator_id}, 其他成員: {other_user_ids}")
        
        # 創建群組
        group_info = await manager.create_group(
            title=group_name,
            about="AI 紅包互動測試群",
            member_user_ids=other_user_ids[:5],  # 先加5個
            is_supergroup=True  # 使用超級群組，支持更多功能
        )
        
        if not group_info:
            logger.error("創建群組失敗")
            return None
        
        self.test_group = group_info
        
        # 邀請剩餘成員
        if len(other_user_ids) > 5:
            remaining = other_user_ids[5:]
            logger.info(f"邀請剩餘成員: {remaining}")
            await manager.invite_users(
                group_info["id"],
                remaining,
                is_channel=True
            )
        
        return group_info
    
    async def all_accounts_join_via_link(self, invite_link: str) -> Dict[int, bool]:
        """所有帳號通過邀請鏈接加入群組"""
        results = {}
        
        for user_id, account in self.accounts.items():
            client: TelegramClient = account["client"]
            
            try:
                from telethon.tl.functions.messages import ImportChatInviteRequest
                
                # 從鏈接提取 hash
                if "+" in invite_link:
                    hash_part = invite_link.split("+")[-1]
                elif "joinchat/" in invite_link:
                    hash_part = invite_link.split("joinchat/")[-1]
                else:
                    # 嘗試直接加入公開群
                    await client.get_entity(invite_link)
                    results[user_id] = True
                    continue
                
                await client(ImportChatInviteRequest(hash_part))
                results[user_id] = True
                logger.info(f"[用戶 {user_id}] 成功加入群組")
                
            except Exception as e:
                error_msg = str(e)
                if "already" in error_msg.lower() or "已經" in error_msg:
                    results[user_id] = True
                    logger.info(f"[用戶 {user_id}] 已在群組中")
                else:
                    results[user_id] = False
                    logger.error(f"[用戶 {user_id}] 加入群組失敗: {e}")
            
            # 間隔避免限流
            await asyncio.sleep(random.uniform(3, 8))
        
        return results
    
    def get_all_user_ids(self) -> List[int]:
        """獲取所有帳號的 user_id"""
        return list(self.accounts.keys())
    
    def get_account_count(self) -> int:
        """獲取帳號數量"""
        return len(self.accounts)


# ==================== 帳號載入（確保每個帳號獨立 API） ====================

def validate_unique_api_credentials(accounts: List[dict]) -> bool:
    """
    驗證所有帳號的 API 憑證是否唯一
    
    Returns:
        True 如果所有帳號都有唯一的 API 憑證
    """
    seen_credentials = set()
    duplicates = []
    
    for account in accounts:
        api_id = account.get("api_id")
        api_hash = account.get("api_hash")
        phone = account.get("phone", "未知")
        
        if not api_id or not api_hash:
            logger.error(f"帳號 {phone} 缺少 API 憑證！")
            return False
        
        credential_key = f"{api_id}:{api_hash}"
        
        if credential_key in seen_credentials:
            duplicates.append(phone)
            logger.error(f"⚠️ 帳號 {phone} 的 API 憑證與其他帳號重複！")
        else:
            seen_credentials.add(credential_key)
    
    if duplicates:
        logger.error("=" * 50)
        logger.error("❌ 發現重複的 API 憑證！")
        logger.error("每個帳號必須使用獨立的 API_ID 和 API_HASH")
        logger.error("請在 my.telegram.org 為每個帳號創建獨立的應用")
        logger.error("=" * 50)
        return False
    
    logger.info(f"✅ 所有 {len(accounts)} 個帳號的 API 憑證都是唯一的")
    return True


def load_accounts_with_validation(excel_path: str) -> List[dict]:
    """
    從 Excel 載入帳號配置並驗證
    
    每個帳號必須有獨立的 api_id 和 api_hash
    """
    import openpyxl
    
    accounts = []
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # 讀取標題行
        headers = [str(cell.value).lower().strip() if cell.value else "" for cell in ws[1]]
        
        # 列名映射
        col_map = {}
        for idx, header in enumerate(headers):
            if header in ["phone", "手机", "手機", "電話", "电话", "mobile"]:
                col_map["phone"] = idx
            elif header in ["api_id", "apiid", "api id"]:
                col_map["api_id"] = idx
            elif header in ["api_hash", "apihash", "api hash"]:
                col_map["api_hash"] = idx
            elif header in ["user_id", "userid", "tg_id", "telegram_id", "id"]:
                col_map["user_id"] = idx
            elif header in ["username", "用户名", "用戶名"]:
                col_map["username"] = idx
            elif header in ["name", "名字", "昵称", "暱稱"]:
                col_map["name"] = idx
            elif header in ["enabled", "启用", "啟用", "active"]:
                col_map["enabled"] = idx
        
        # 驗證必需列
        if "phone" not in col_map:
            logger.error("Excel 缺少 phone 列！")
            return []
        if "api_id" not in col_map:
            logger.error("Excel 缺少 api_id 列！")
            return []
        if "api_hash" not in col_map:
            logger.error("Excel 缺少 api_hash 列！")
            return []
        
        # 讀取數據行
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
            
            account = {"row": row_idx}
            
            # 電話（必需）
            phone = row[col_map["phone"]]
            if not phone:
                continue
            account["phone"] = str(phone).strip()
            
            # API ID（必需）
            api_id = row[col_map["api_id"]]
            if not api_id:
                logger.warning(f"第 {row_idx} 行缺少 api_id")
                continue
            account["api_id"] = int(api_id)
            
            # API Hash（必需）
            api_hash = row[col_map["api_hash"]]
            if not api_hash:
                logger.warning(f"第 {row_idx} 行缺少 api_hash")
                continue
            account["api_hash"] = str(api_hash).strip()
            
            # 可選欄位
            if "user_id" in col_map and row[col_map["user_id"]]:
                account["user_id"] = int(row[col_map["user_id"]])
            
            if "username" in col_map and row[col_map["username"]]:
                account["username"] = str(row[col_map["username"]]).strip()
            
            if "name" in col_map and row[col_map["name"]]:
                account["name"] = str(row[col_map["name"]]).strip()
            
            # 啟用狀態
            if "enabled" in col_map:
                enabled = row[col_map["enabled"]]
                account["enabled"] = enabled in [1, "1", True, "true", "True", "yes", "Yes", None, ""]
            else:
                account["enabled"] = True
            
            if account.get("enabled", True):
                accounts.append(account)
        
        logger.info(f"從 Excel 載入了 {len(accounts)} 個帳號配置")
        
        # 驗證 API 憑證唯一性
        if not validate_unique_api_credentials(accounts):
            logger.error("API 憑證驗證失敗，請修復後重試")
            return []
        
        return accounts
        
    except Exception as e:
        logger.error(f"讀取 Excel 失敗: {e}")
        return []


if __name__ == "__main__":
    # 測試載入
    logging.basicConfig(level=logging.INFO)
    
    import sys
    if len(sys.argv) > 1:
        accounts = load_accounts_with_validation(sys.argv[1])
        for acc in accounts:
            print(f"Phone: {acc['phone']}, API_ID: {acc['api_id']}, Hash: {acc['api_hash'][:8]}...")
